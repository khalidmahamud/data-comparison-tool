import re
import pandas as pd
from pathlib import Path
from typing import Optional, Tuple
from openpyxl import load_workbook
from prompt import inject_variables, read_file
from ai import ask
from config import config


def extract_standard_letters(text):
    pattern = r'[^\w\s]'
    
    result = re.sub(pattern, '', text, flags=re.UNICODE)
    
    return result


def read_row(row_idx: int, input_file: str) -> Tuple[str, str, str]:
    excel_path = Path(input_file)
    if not excel_path.exists():
        raise FileNotFoundError(f"Input file '{excel_path}' not found")
    
    # Load the Excel file
    try:
        df = pd.read_excel(excel_path, sheet_name='hadith')
    except Exception as e:
        raise ValueError(f"Error reading Excel file: {str(e)}")
    
    # Validate row index
    if row_idx < 0 or row_idx >= len(df):
        raise ValueError(f"Row index {row_idx} out of bounds (0-{len(df)-1})")
    
    # Get the required data
    hadith_details = df.loc[row_idx, 'hadith_details'] if 'hadith_details' in df.columns else ""
    arabic_text = df.loc[row_idx, 'arabic_text'] if 'arabic_text' in df.columns else ""
    current_analysis = df.loc[row_idx, 'analysis-3'] if 'analysis-3' in df.columns else ""
    
    return arabic_text, hadith_details, current_analysis

def save_to_excel(row_idx: int, new_text: str, input_file: str, output_file: str = None) -> bool:
    """
    Save the generated text back to Excel file.
    
    Args:
        row_idx (int): The row index to update
        new_text (str): The new text to save
        input_file (str): Path to input Excel file
        output_file (str): Optional path to output file. If not provided, updates input file
        
    Returns:
        bool: True if save was successful, False otherwise
    """
    excel_path = Path(input_file)
    output_path = Path(output_file) if output_file else excel_path
    
    try:
        # Try with openpyxl first to preserve formatting
        wb = load_workbook(excel_path)
        if 'hadith' not in wb.sheetnames:
            raise ValueError("'hadith' sheet not found in Excel file")
        
        ws = wb['hadith']
        
        # Find the analysis-3 column index
        analysis3_col_idx = None
        for idx, cell in enumerate(next(ws.rows)):
            if cell.value == 'analysis-3':
                analysis3_col_idx = idx
                break
        
        if analysis3_col_idx is None:
            analysis3_col_idx = 6  # Default to column G
        
        # Calculate Excel row (add 2 to account for 0-based index and header row)
        excel_row = row_idx + 2
        col_letter = chr(65 + analysis3_col_idx)
        cell_address = f'{col_letter}{excel_row}'
        
        # Update the cell
        ws[cell_address].value = new_text
        wb.save(output_path)
        return True
        
    except Exception as e:
        print(f"Error with openpyxl: {str(e)}")
        
        # Fallback to pandas
        try:
            df = pd.read_excel(excel_path, sheet_name='hadith')
            
            if 'analysis-3' in df.columns:
                df.loc[row_idx, 'analysis-3'] = new_text
            else:
                if len(df.columns) <= 6:
                    # Add columns if needed
                    while len(df.columns) < 7:
                        df[f"Column_{len(df.columns)}"] = ""
                df.iloc[row_idx, 6] = new_text
            
            df.to_excel(output_path, sheet_name='hadith', index=False)
            return True
            
        except Exception as e2:
            print(f"Pandas fallback failed: {str(e2)}")
            return False


def generate(row_idx: int) -> str:
    arabic_text, hadith_details, current_analysis = read_row(row_idx, config.file_settings.input_file)

    query = inject_variables(read_file("./prompts/regenerate_hadis_prompt.md"), {
        "hadis_arabic_text": arabic_text,
        "hadis_translated_bangla": hadith_details,
        "current_analysis": current_analysis
    })


    return ask(query).text
    