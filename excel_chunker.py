#!/usr/bin/env python3
import os
import re
import sys
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, Protection
from copy import copy
from config import config

def copy_cell_style(source_cell, target_cell):
    """
    Copy all styling properties from source cell to target cell
    """
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def validate_columns(worksheet):
    """
    Validate that required columns exist in the worksheet.
    
    Args:
        worksheet: The openpyxl worksheet to validate
        
    Returns:
        tuple: (is_valid, missing_columns, has_ratio)
    """
    # Get column names from config
    primary_text_col = config.excel_settings.columns.get('primary_text', 'hadith_details')
    secondary_text_col = config.excel_settings.columns.get('secondary_text', 'analysis-3')
    ratio_col = config.excel_settings.columns.get('ratio', 'ratio')
    
    # Extract header row values
    header_values = [cell.value for cell in next(worksheet.rows)]
    
    # Check if required columns exist
    missing_columns = []
    if primary_text_col not in header_values:
        missing_columns.append(primary_text_col)
    if secondary_text_col not in header_values:
        missing_columns.append(secondary_text_col)
        
    # Ratio column is optional
    has_ratio = ratio_col in header_values
    
    is_valid = len(missing_columns) == 0
    
    return is_valid, missing_columns, has_ratio


def split_excel(input_file, output_dir='chunks', rows_per_chunk=500):
    """
    Split a large Excel file into smaller chunks with style preservation.
    
    Args:
        input_file (str): Path to the input Excel file
        output_dir (str): Directory to save the chunks
        rows_per_chunk (int): Maximum number of rows per chunk
    
    Returns:
        list: List of generated chunk files
    """
    start_time = time.time()
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Starting Excel file splitting process")
    
    # Ensure output directory exists
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Load the workbook once
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Loading workbook: {input_file}")
    wb = load_workbook(input_file)
    load_time = time.time()
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Workbook loaded in {load_time - start_time:.2f} seconds")
    
    # Get the sheet from configuration
    sheet_name = config.excel_settings.sheet_name
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Warning: Configured sheet '{config.excel_settings.sheet_name}' not found. Using '{sheet_name}' instead.")
    
    ws = wb[sheet_name]
    
    # Validate columns
    is_valid, missing_columns, has_ratio = validate_columns(ws)
    if not is_valid:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Error: Missing required columns: {', '.join(missing_columns)}")
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Required columns: {config.excel_settings.columns.get('primary_text', 'hadith_details')}, {config.excel_settings.columns.get('secondary_text', 'analysis-3')}")
        return []
    
    if not has_ratio:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Warning: Optional ratio column '{config.excel_settings.columns.get('ratio', 'ratio')}' not found. It will be created during processing.")
    
    # Get total rows (excluding header)
    total_rows = ws.max_row - 1  # Subtract 1 for header row
    
    # Calculate number of chunks needed
    num_chunks = (total_rows + rows_per_chunk - 1) // rows_per_chunk
    
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Processing sheet: {sheet_name}")
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Total rows: {total_rows}")
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Rows per chunk: {rows_per_chunk}")
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Number of chunks: {num_chunks}")
    
    chunk_files = []
    
    # Process each chunk
    for chunk_idx in range(num_chunks):
        chunk_start_time = time.time()
        # Calculate row range for this chunk
        start_row = chunk_idx * rows_per_chunk + 2  # +2 because row 1 is header, and we want to start from row 2
        end_row = min((chunk_idx + 1) * rows_per_chunk + 1, total_rows + 1)  # +1 because row 1 is the header
        
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Creating chunk {chunk_idx+1}/{num_chunks} with rows {start_row-1}-{end_row-1}")
        
        # Create a new workbook for the chunk
        chunk_wb = Workbook()
        chunk_ws = chunk_wb.active
        chunk_ws.title = sheet_name
        
        # Copy column dimensions
        for col_letter, column_dimension in ws.column_dimensions.items():
            chunk_ws.column_dimensions[col_letter].width = column_dimension.width
            chunk_ws.column_dimensions[col_letter].hidden = column_dimension.hidden
        
        # Add header row first
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Copying header row with styles")
        for col_idx in range(1, ws.max_column + 1):
            source_cell = ws.cell(row=1, column=col_idx)
            target_cell = chunk_ws.cell(row=1, column=col_idx)
            
            # Copy value and style
            target_cell.value = source_cell.value
            copy_cell_style(source_cell, target_cell)
        
        # Copy only the rows needed for this chunk with styles
        row_count = end_row - start_row + 1
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Copying {row_count} data rows with styles")
        
        # Show progress for large chunks
        progress_interval = max(1, row_count // 10)
        
        for i, (dest_row, source_row) in enumerate(zip(range(2, row_count + 2), range(start_row, end_row + 1))):
            # Show progress for large chunks
            if i % progress_interval == 0 and i > 0:
                percent_done = (i / row_count) * 100
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Progress: {percent_done:.1f}% ({i}/{row_count} rows)")
            
            # Also copy row dimensions
            if source_row in ws.row_dimensions:
                chunk_ws.row_dimensions[dest_row].height = ws.row_dimensions[source_row].height
                chunk_ws.row_dimensions[dest_row].hidden = ws.row_dimensions[source_row].hidden
                
            for col_idx in range(1, ws.max_column + 1):
                source_cell = ws.cell(row=source_row, column=col_idx)
                target_cell = chunk_ws.cell(row=dest_row, column=col_idx)
                
                # Copy value and style
                target_cell.value = source_cell.value
                copy_cell_style(source_cell, target_cell)
        
        # Define chunk filename
        chunk_filename = f"chunk_{chunk_idx+1}_rows_{start_row-1}-{end_row-1}.xlsx"
        chunk_path = os.path.join(output_dir, chunk_filename)
        
        # Save the chunk
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Saving chunk {chunk_idx+1}/{num_chunks}: {chunk_path}")
        chunk_wb.save(chunk_path)
        chunk_files.append(chunk_path)
        
        chunk_end_time = time.time()
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Chunk {chunk_idx+1} completed in {chunk_end_time - chunk_start_time:.2f} seconds")
    
    end_time = time.time()
    total_time = end_time - start_time
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Splitting complete. Created {len(chunk_files)} chunks in {total_time:.2f} seconds")
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Average time per chunk: {total_time/len(chunk_files):.2f} seconds")
    return chunk_files

def merge_excel(chunk_dir='chunks', output_file=None):
    """
    Merge chunked Excel files back into a single file with style preservation.
    
    Args:
        chunk_dir (str): Directory containing the chunk files
        output_file (str): Output file path. If None, will be 'merged_output.xlsx'
    
    Returns:
        str: Path to the merged file
    """
    start_time = time.time()
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Starting Excel file merging process")
    
    # Find all chunk files
    chunk_files = []
    chunk_pattern = re.compile(r'chunk_(\d+)_rows_(\d+)-(\d+)\.xlsx')
    
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Scanning directory for chunk files: {chunk_dir}")
    for filename in os.listdir(chunk_dir):
        if chunk_pattern.match(filename):
            chunk_match = chunk_pattern.match(filename)
            chunk_num = int(chunk_match.group(1))
            start_row = int(chunk_match.group(2))
            end_row = int(chunk_match.group(3))
            
            chunk_files.append({
                'filename': os.path.join(chunk_dir, filename),
                'chunk_num': chunk_num,
                'start_row': start_row,
                'end_row': end_row,
                'row_count': end_row - start_row + 1
            })
    
    # Sort by chunk number
    chunk_files.sort(key=lambda x: x['chunk_num'])
    
    if not chunk_files:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] No chunk files found")
        return None
    
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Found {len(chunk_files)} chunk files to merge")
    
    # Calculate total rows for progress reporting
    total_rows = sum(chunk['row_count'] for chunk in chunk_files)
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Total rows to process: {total_rows}")
    
    # Load the first chunk as our base workbook
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Loading first chunk: {chunk_files[0]['filename']}")
    first_chunk = load_workbook(chunk_files[0]['filename'])
    
    # Get the sheet from configuration
    sheet_name = config.excel_settings.sheet_name
    if sheet_name not in first_chunk.sheetnames:
        sheet_name = first_chunk.sheetnames[0]
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Warning: Configured sheet '{config.excel_settings.sheet_name}' not found in first chunk. Using '{sheet_name}' instead.")
    
    first_chunk_ws = first_chunk[sheet_name]
    
    # Validate first chunk columns
    is_valid, missing_columns, has_ratio = validate_columns(first_chunk_ws)
    if not is_valid:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Error: Missing required columns in first chunk: {', '.join(missing_columns)}")
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Required columns: {config.excel_settings.columns.get('primary_text', 'hadith_details')}, {config.excel_settings.columns.get('secondary_text', 'analysis-3')}")
        return None
    
    if not has_ratio:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Warning: Optional ratio column '{config.excel_settings.columns.get('ratio', 'ratio')}' not found in first chunk. It will need to be calculated after merge.")
    
    # Create a new workbook for the merged data
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = sheet_name
    
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Copying column dimensions and styles")
    # Copy column dimensions from first chunk
    for col_letter, column_dimension in first_chunk_ws.column_dimensions.items():
        merged_ws.column_dimensions[col_letter].width = column_dimension.width
        merged_ws.column_dimensions[col_letter].hidden = column_dimension.hidden
    
    # Current row in the merged worksheet (start at 1)
    current_row = 1
    rows_processed = 0
    last_progress_report = 0
    
    # Process all chunks
    for i, chunk_info in enumerate(chunk_files):
        chunk_start_time = time.time()
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Processing chunk {i+1}/{len(chunk_files)}: {chunk_info['filename']}")
        
        # Load the chunk
        chunk_wb = load_workbook(chunk_info['filename'])
        
        # Verify the sheet exists in this chunk
        if sheet_name not in chunk_wb.sheetnames:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Warning: Sheet '{sheet_name}' not found in chunk {i+1}. Skipping chunk.")
            continue
            
        chunk_ws = chunk_wb[sheet_name]
        
        # Validate this chunk's columns (only for chunks after the first one)
        if i > 0:
            is_valid, missing_columns, _ = validate_columns(chunk_ws)
            if not is_valid:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Warning: Missing required columns in chunk {i+1}: {', '.join(missing_columns)}. Skipping chunk.")
                continue
        
        # For the first chunk, include all rows (including header)
        start_row = 1 if i == 0 else 2  # Skip header for all but first chunk
        
        rows_in_chunk = chunk_ws.max_row - (start_row - 1)
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Processing {rows_in_chunk} rows from chunk {i+1}")
        
        # Copy rows from chunk to merged workbook (with or without header based on chunk number)
        for row_idx in range(start_row, chunk_ws.max_row + 1):
            rows_processed += 1
            
            # Report progress every 10% of total rows
            progress_percent = (rows_processed / total_rows) * 100
            if progress_percent - last_progress_report >= 10:
                print(f"[{datetime.now().strftime('%H:%M:%S')}] Merge progress: {progress_percent:.1f}% ({rows_processed}/{total_rows} rows)")
                last_progress_report = progress_percent // 10 * 10
            
            # Copy row dimensions
            if row_idx in chunk_ws.row_dimensions:
                merged_ws.row_dimensions[current_row].height = chunk_ws.row_dimensions[row_idx].height
                merged_ws.row_dimensions[current_row].hidden = chunk_ws.row_dimensions[row_idx].hidden
            
            # Copy cell values and styles
            for col_idx in range(1, chunk_ws.max_column + 1):
                source_cell = chunk_ws.cell(row=row_idx, column=col_idx)
                target_cell = merged_ws.cell(row=current_row, column=col_idx)
                
                # Copy value and style
                target_cell.value = source_cell.value
                copy_cell_style(source_cell, target_cell)
            
            current_row += 1
        
        chunk_end_time = time.time()
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Chunk {i+1} processed in {chunk_end_time - chunk_start_time:.2f} seconds")
    
    # Save the merged workbook
    if output_file is None:
        output_file = config.file_settings.merged_file
    
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Saving merged file: {output_file}")
    merged_wb.save(output_file)
    
    end_time = time.time()
    total_time = end_time - start_time
    
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Merge complete. Total rows: {current_row - 1}")
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Merging completed in {total_time:.2f} seconds")
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Average time per row: {total_time/(current_row-1):.4f} seconds")
    
    # Inform if ratio column was missing
    if not has_ratio:
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Note: The ratio column was not present in the chunks. You may need to recalculate ratios.")
    
    return output_file

def main():
    # Use values directly from config file instead of command-line arguments
    input_file = config.file_settings.input_file
    chunk_dir = config.file_settings.chunks_directory
    output_file = config.file_settings.merged_file
    rows_per_chunk = 500  # Default value
    
    # Check if processing config has rows_per_chunk setting
    if hasattr(config.file_settings, 'rows_per_chunk'):
        rows_per_chunk = config.file_settings.rows_per_chunk
    
    # Get action from config file
    action = "split"  # Default action
    if hasattr(config.file_settings, 'action'):
        action = config.file_settings.action.lower()
    
    # Validate action
    if action not in ['split', 'merge']:
        print(f"Error: Invalid action '{action}' in config. Choose 'split' or 'merge'")
        sys.exit(1)
    
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Running with action: {action}")
    
    if action == 'split':
        split_excel(input_file, chunk_dir, rows_per_chunk)
    elif action == 'merge':
        merge_excel(chunk_dir, output_file)

if __name__ == '__main__':
    main()
