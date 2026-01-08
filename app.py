from flask import Flask, render_template, request, jsonify, redirect, url_for, send_from_directory
import pandas as pd, math, os, difflib, re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from urllib.parse import urlencode
import uuid
from datetime import datetime
import threading
import time
import shutil
from werkzeug.utils import secure_filename

from pathlib import Path
from src.prompt import inject_variables
from src.ai import ask
from src.generate_cell import generate, extract_standard_letters
from src.config import config, load_config, ServerConfig

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = ServerConfig.get_secret_key()
app.config['MAX_CONTENT_LENGTH'] = ServerConfig.get_max_upload_size()

# Initialize database
try:
    from src.database import init_db
    init_db(app)
except Exception as e:
    print(f"Warning: Database initialization failed: {e}")

# Configure upload folder
UPLOAD_FOLDER = ServerConfig.get_upload_folder()
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
Path(UPLOAD_FOLDER).mkdir(parents=True, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    """Check if file has allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# --- File Safety and Caching System ---
file_lock = threading.RLock()
excel_cache = {'df': None, 'mtime': None, 'path': None, 'color_status': None, 'color_mtime': None}

def safe_load_workbook(input_file, read_only=False, max_retries=3, retry_delay=0.1):
    """Safely load workbook with retries and file validation"""
    for attempt in range(max_retries):
        try:
            # Check if file exists and is readable
            if not input_file or not os.path.exists(input_file):
                raise FileNotFoundError(f"File not found: {input_file}")
            
            # Check file size (basic validation)
            if os.path.getsize(input_file) == 0:
                raise ValueError(f"File is empty: {input_file}")
            
            # Try to load the workbook
            wb = load_workbook(input_file, read_only=read_only, data_only=True)
            return wb
            
        except Exception as e:
            if attempt < max_retries - 1:
                print(f"Attempt {attempt + 1} failed loading {input_file}: {e}. Retrying in {retry_delay}s...")
                time.sleep(retry_delay)
                retry_delay *= 2  # Exponential backoff
            else:
                print(f"Failed to load {input_file} after {max_retries} attempts: {e}")
                raise

def safe_save_workbook(wb, input_file, max_retries=3, retry_delay=0.1):
    """Safely save workbook with backup and retries"""
    backup_file = f"{input_file}.backup"
    
    for attempt in range(max_retries):
        try:
            # Create backup if original exists
            if os.path.exists(input_file):
                shutil.copy2(input_file, backup_file)
            
            # Save the workbook
            wb.save(input_file)
            
            # Remove backup if save was successful
            if os.path.exists(backup_file):
                os.remove(backup_file)
            
            # Clear cache after successful save
            with file_lock:
                excel_cache['df'] = None
                excel_cache['color_status'] = None
            
            return True
            
        except Exception as e:
            if attempt < max_retries - 1:
                print(f"Save attempt {attempt + 1} failed for {input_file}: {e}. Retrying in {retry_delay}s...")
                time.sleep(retry_delay)
                retry_delay *= 2
            else:
                # Restore backup if save failed
                if os.path.exists(backup_file) and os.path.exists(input_file):
                    try:
                        shutil.copy2(backup_file, input_file)
                        print(f"Restored backup for {input_file}")
                    except Exception as restore_error:
                        print(f"Failed to restore backup: {restore_error}")
                
                print(f"Failed to save {input_file} after {max_retries} attempts: {e}")
                raise

def get_cached_dataframe(input_file, sheet_name):
    """Get cached DataFrame or load from file if cache is stale"""
    with file_lock:
        try:
            current_mtime = os.path.getmtime(input_file)

            # Check if cache is valid (same file, mtime, AND sheet_name)
            if (excel_cache['df'] is not None and
                excel_cache['mtime'] == current_mtime and
                excel_cache['path'] == input_file and
                excel_cache.get('sheet_name') == sheet_name):
                return excel_cache['df'].copy()

            # Load fresh data
            print(f"Loading fresh data from {input_file}, sheet: {sheet_name}")
            df = pd.read_excel(input_file, sheet_name=sheet_name, engine='openpyxl')

            # Update cache
            excel_cache['df'] = df.copy()
            excel_cache['mtime'] = current_mtime
            excel_cache['path'] = input_file
            excel_cache['sheet_name'] = sheet_name

            return df.copy()

        except Exception as e:
            print(f"Error loading DataFrame: {e}")
            # Fallback to direct load
            return pd.read_excel(input_file, sheet_name=sheet_name, engine='openpyxl')

def get_cached_color_status(input_file):
    """Get cached color status or load from file if cache is stale"""
    if not input_file:
        return {}
    with file_lock:
        try:
            current_mtime = os.path.getmtime(input_file)
            
            if (excel_cache['color_status'] is not None and 
                excel_cache['color_mtime'] == current_mtime and 
                excel_cache['path'] == input_file):
                return excel_cache['color_status'].copy()
            
            # Load fresh color status
            print(f"Loading fresh color status from {input_file}")
            color_status = _load_color_status(input_file)
            
            # Update cache
            excel_cache['color_status'] = color_status.copy()
            excel_cache['color_mtime'] = current_mtime
            
            return color_status.copy()
            
        except Exception as e:
            print(f"Error loading color status: {e}")
            return {}

def _load_color_status(input_file):
    """Internal function to load color status from file"""
    if not input_file or not os.path.exists(input_file): 
        return {}
    
    try:
        wb = safe_load_workbook(input_file, read_only=True)
    except Exception:
        return {}
    
    sheet_name = get_sheet_name()
    if sheet_name not in wb.sheetnames: 
        return {}
    ws = wb[sheet_name]
    
    primary_text_col_name = get_column_name('primary_text')
    secondary_text_col_name = get_column_name('secondary_text')
    
    primary_text_col_idx = secondary_text_col_idx = None
    try:
        header_row = next(ws.rows)
        for idx, cell in enumerate(header_row):
            col_name = cell.value
            if col_name == primary_text_col_name: 
                primary_text_col_idx = idx
            elif col_name == secondary_text_col_name: 
                secondary_text_col_idx = idx
    except StopIteration:
        return {}
    
    primary_text_col_idx = 0 if primary_text_col_idx is None else primary_text_col_idx
    secondary_text_col_idx = 1 if secondary_text_col_idx is None else secondary_text_col_idx
    
    color_status = {}
    
    def check_cell_color(cell, row_dict, col_key):
        if not (hasattr(cell, 'fill') and cell.fill and cell.fill.fill_type != 'none'): 
            return
        if not (hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb): 
            return
        
        rgb = cell.fill.start_color.rgb
        rgb_str = str(rgb).upper()
        if not (rgb_str and rgb_str != "00000000" and not rgb_str.endswith("000000")): 
            return
        
        row_dict[f'col_{col_key}'] = True
        
        if "FF0000" in rgb_str or "FFFF0000" in rgb_str or rgb_str.endswith("FF0000"): 
            row_dict[f'col_{col_key}_type'] = 'red'
        elif "00FF00" in rgb_str or rgb_str == "FF00FF00": 
            row_dict[f'col_{col_key}_type'] = 'green'
        elif "FFFF00" in rgb_str or rgb_str == "FFFFFF00": 
            row_dict[f'col_{col_key}_type'] = 'yellow'
        elif rgb_str and len(rgb_str) >= 6:
            rgb_part = rgb_str[-6:] if len(rgb_str) > 6 else rgb_str
            r_val, g_val, b_val = rgb_part[0:2], rgb_part[2:4], rgb_part[4:6]
            row_dict[f'col_{col_key}_type'] = 'red' if r_val in ["FF", "F0", "E0"] and g_val in ["00", "10", "20", "30"] and b_val in ["00", "10", "20", "30"] else 'green'

    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if len(row) > max(primary_text_col_idx, secondary_text_col_idx):
                col_a_cell, col_b_cell = row[primary_text_col_idx], row[secondary_text_col_idx]
                excel_row_idx = row_idx
                color_status[excel_row_idx] = {'col_a': False, 'col_b': False, 'col_a_type': None, 'col_b_type': None}
                check_cell_color(col_a_cell, color_status[excel_row_idx], 'a')
                check_cell_color(col_b_cell, color_status[excel_row_idx], 'b')
    except Exception as e:
        print(f"Error reading color status: {e}")
    
    try:
        wb.close()
    except:
        pass
    
    return color_status

def batch_update_excel_cells(input_file, generated_texts, clear_fill=True):
    """
    Safely update multiple Excel cells in a single operation
    
    Args:
        input_file: Path to the Excel file
        generated_texts: Dict mapping row_idx to new_text
        clear_fill: Whether to clear cell formatting
    
    Returns:
        List of result dicts for each updated row
    """
    results = []
    
    with file_lock:
        wb = safe_load_workbook(input_file)
        sheet_name = get_sheet_name()
        
        if sheet_name not in wb.sheetnames:
            raise ValueError(f'{sheet_name} sheet not found')
        
        ws = wb[sheet_name]
        
        # Find column indices
        header = next(ws.rows)
        primary_text_col_name = get_column_name('primary_text')
        secondary_text_col_name = get_column_name('secondary_text')
        
        secondary_text_col_idx = 1
        primary_text_col_idx = 0
        
        for idx, cell in enumerate(header):
            if cell.value == secondary_text_col_name:
                secondary_text_col_idx = idx
            elif cell.value == primary_text_col_name:
                primary_text_col_idx = idx
        
        # Update all cells at once
        for row_idx, new_text in generated_texts.items():
            excel_row = row_idx + 2
            cell_address = f'{get_column_letter(secondary_text_col_idx + 1)}{excel_row}'
            ws[cell_address].value = new_text
            if clear_fill:
                ws[cell_address].fill = PatternFill(fill_type=None)
        
        # Save the workbook once after all updates
        safe_save_workbook(wb, input_file)
        
        # Now collect all comparison results
        for row_idx, new_text in generated_texts.items():
            excel_row = row_idx + 2
            
            # Get original text for comparison
            col_a_cell = ws[f'{get_column_letter(primary_text_col_idx + 1)}{excel_row}']
            col_a_text = str(col_a_cell.value) if col_a_cell.value is not None else ''
            highlighted_a, highlighted_b, status = compare_text(col_a_text, new_text)
            
            # Fetch color status for this row
            color_status = get_cell_color_status()
            row_approval = color_status.get(excel_row, {'col_b': False, 'col_b_type': None})
            col_b_approved = row_approval['col_b']
            col_b_type = row_approval['col_b_type']
            
            results.append({
                'status': 'success',
                'row_idx': row_idx,
                'new_text': new_text,
                'highlighted_html': highlighted_b,
                'highlighted_a_html': highlighted_a,
                'diff_status': status,
                'col_b_approved': col_b_approved,
                'col_b_type': col_b_type
            })
    
    return results

# --- Configuration Loading ---
# Global variable to track the currently selected chunk
current_chunk = None

# Function to get all available chunks
def get_available_chunks():
    chunks_dir = 'chunks'
    if not os.path.exists(chunks_dir):
        return []
    
    chunks = []
    chunk_pattern = re.compile(r'chunk_(\d+)_rows_(\d+)-(\d+)\.xlsx')
    
    for filename in os.listdir(chunks_dir):
        if chunk_pattern.match(filename):
            chunk_match = chunk_pattern.match(filename)
            chunk_num = int(chunk_match.group(1))
            start_row = int(chunk_match.group(2))
            end_row = int(chunk_match.group(3))
            
            chunks.append({
                'filename': os.path.join(chunks_dir, filename),
                'chunk_num': chunk_num,
                'display_name': f"C{chunk_num} ({start_row}-{end_row})",
                'start_row': start_row,
                'end_row': end_row
            })
    
    # Sort by chunk number
    chunks.sort(key=lambda x: x['chunk_num'])
    return chunks

# Initialize current_chunk to the first available chunk
chunks = get_available_chunks()
if chunks:
    current_chunk = chunks[0]['filename']

def reload_config(config_path: str = 'config_flash.yaml'):
    # Reloads the configuration from a YAML file.
    global config
    try:
        config = load_config(config_path)
    except Exception as e:
        print(f"Error loading configuration: {e}")

def get_input_file_path():
    """Get the currently selected file path, or None if no file is selected."""
    global current_chunk
    if current_chunk and os.path.exists(current_chunk):
        return current_chunk
    return None

def get_uploaded_files_list():
    """Get list of uploaded files with metadata."""
    files = []
    upload_folder = app.config['UPLOAD_FOLDER']

    if os.path.exists(upload_folder):
        for filename in os.listdir(upload_folder):
            if allowed_file(filename):
                filepath = os.path.join(upload_folder, filename)
                stat = os.stat(filepath)
                files.append({
                    'filename': filename,
                    'filepath': filepath,
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    'display_size': f"{stat.st_size / 1024:.1f} KB" if stat.st_size < 1024*1024 else f"{stat.st_size / (1024*1024):.1f} MB"
                })

    # Sort by modified date (newest first)
    files.sort(key=lambda x: x['modified'], reverse=True)
    return files

def get_sheet_name() -> str:
    # Get the sheet name from configuration
    try:
        return config.excel_settings.sheet_name
    except Exception as e:
        print(f"Error accessing sheet name from configuration: {e}. Using default.")
        return 'hadith'

def get_column_name(column_key: str) -> str:
    # Get the column name for a specific key from configuration
    try:
        return config.excel_settings.columns.get(column_key, column_key)
    except Exception as e:
        # Default fallbacks for essential columns
        defaults = {
            'primary_text': 'hadith_details',
            'secondary_text': 'analysis-3',
            'ratio': 'ratio',
            'number': 'number',
            'arabic_text': 'arabic_text'
        }
        print(f"Error accessing column name from configuration: {e}. Using default.")
        return defaults.get(column_key, column_key)

# Ensure configuration is loaded when the app starts
reload_config()
# --- End Configuration Loading ---

def compare_text(text1, text2):
    # Handle case where inputs might be Series (e.g., from duplicate columns)
    if isinstance(text1, pd.Series):
        text1 = text1.iloc[0] if len(text1) > 0 else None
    if isinstance(text2, pd.Series):
        text2 = text2.iloc[0] if len(text2) > 0 else None

    if pd.isna(text1) and pd.isna(text2): return "", "", "same"
    elif pd.isna(text1): 
        replaced_text = str(text2).replace("\n", "<br>")
        return "", f'<span class="added">{replaced_text}</span>', "different"
    elif pd.isna(text2): 
        replaced_text = str(text1).replace("\n", "<br>")
        return f'<span class="removed">{replaced_text}</span>', "", "different"
    
    text1, text2 = str(text1), str(text2)
    if text1 == text2: return text1.replace("\n", "<br>"), text2.replace("\n", "<br>"), "same"
    
    line_break_marker = " ¶ "
    text1_prep = text1.replace('\r\n', '\n').replace('\r', '\n').replace('\n', line_break_marker)
    text2_prep = text2.replace('\r\n', '\n').replace('\r', '\n').replace('\n', line_break_marker)
    
    words1 = re.split(r'(\s+)', text1_prep)
    words2 = re.split(r'(\s+)', text2_prep)
    words1 = [word for word in words1 if word]
    words2 = [word for word in words2 if word]
    
    matcher = difflib.SequenceMatcher(None, words1, words2, autojunk=False)



    

    result1, result2 = [], []
    diff_id_counter = 0
    
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        words1_segment = "".join(words1[i1:i2])
        words2_segment = "".join(words2[j1:j2])
        
        if tag == 'equal':
            result1.append(words1_segment)
            result2.append(words2_segment)
        else:
            # Create deterministic diff_id based on position and content
            diff_id = f"diff-{diff_id_counter}-{tag}-{i1}-{i2}-{j1}-{j2}"
            diff_id_counter += 1
            if words1_segment:
                result1.append(f'<span class="removed" data-diff-id="{diff_id}">{words1_segment}</span>')
            if words2_segment:
                result2.append(f'<span class="added" data-diff-id="{diff_id}">{words2_segment}</span>')
    
    final_text1 = "".join(result1).replace(line_break_marker.strip(), "<br>")
    final_text2 = "".join(result2).replace(line_break_marker.strip(), "<br>")
    
    final_text1 = re.sub(r'<span class="(?:added|removed)" data-diff-id="[^"]*"></span>', '', final_text1)
    final_text2 = re.sub(r'<span class="(?:added|removed)" data-diff-id="[^"]*"></span>', '', final_text2)
    
    return final_text1, final_text2, "different"

def get_cell_color_status():
    input_file = get_input_file_path()
    return get_cached_color_status(input_file)

def get_excel_data(rows_per_page=10, page=1, filter_change_enabled=False, filter_change_value=None, filter_change_lt_value=None, filter_change_from_value=None, filter_change_to_value=None, filter_color_a='any', filter_color_b='any', sort_order='asc', filter_id=None, filter_comment=None):
    input_file = get_input_file_path()
    if not input_file or not os.path.exists(input_file):
        return [], 0, 0, False

    change_col_exists = False
    try:
        sheet_name = get_sheet_name()
        df = get_cached_dataframe(input_file, sheet_name)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return [], 0, 0, False

    primary_text_col = get_column_name('primary_text')
    secondary_text_col = get_column_name('secondary_text')
    ratio_col = get_column_name('ratio')
    number_col = get_column_name('number')

    if primary_text_col not in df.columns or secondary_text_col not in df.columns:
        if len(df.columns) >= 2: 
            df = df.rename(columns={df.columns[0]: primary_text_col, df.columns[1]: secondary_text_col})
        else: 
            return [], 0, 0, False

    # Only calculate ratios if column doesn't exist
    if ratio_col not in df.columns:
        print("Calculating ratios for DataFrame...")
        df[ratio_col] = df.apply(lambda row: difflib.SequenceMatcher(
            None, 
            str(row[primary_text_col]) if pd.notna(row[primary_text_col]) else "",
            str(row[secondary_text_col]) if pd.notna(row[secondary_text_col]) else "",
            autojunk=False
        ).ratio() * 100, axis=1)

        # Save the ratio column back to Excel file in a separate thread to avoid blocking
        def save_ratio_async():
            try:
                with file_lock:
                    wb = safe_load_workbook(input_file)
                    sheet_name = get_sheet_name()
                    if sheet_name not in wb.sheetnames:
                        print(f"Warning: '{sheet_name}' sheet not found in Excel file")
                        return
                    
                    ws = wb[sheet_name]
                    
                    # Find the last column index or existing ratio column
                    ratio_col_idx = None
                    header_row = next(ws.rows)
                    for idx, cell in enumerate(header_row):
                        if cell.value == ratio_col:
                            ratio_col_idx = idx
                            break
                    
                    if ratio_col_idx is None:
                        last_col_idx = len(list(header_row))
                        ratio_col_letter = get_column_letter(last_col_idx + 1)
                        ws[f'{ratio_col_letter}1'] = ratio_col
                        ratio_col_idx = last_col_idx
                    
                    ratio_col_letter = get_column_letter(ratio_col_idx + 1)
                    
                    # Add ratio values for each row
                    for idx, ratio in enumerate(df[ratio_col], start=2):
                        ws[f'{ratio_col_letter}{idx}'] = ratio
                    
                    safe_save_workbook(wb, input_file)
                    
            except Exception as e:
                print(f"Warning: Could not save ratio column to Excel file: {e}")
        
        # Run ratio saving in background
        import threading
        threading.Thread(target=save_ratio_async, daemon=True).start()

    number_col_exists = number_col in df.columns

    if 'change' in df.columns:
        change_col_exists = True
        df['change'] = pd.to_numeric(df['change'], errors='coerce')

    # Sort by ratio based on sort_order parameter
    if sort_order == 'asc':
        df = df.sort_values(by=ratio_col, ascending=True)
    elif sort_order == 'desc':
        df = df.sort_values(by=ratio_col, ascending=False)

    # Apply filters (keeping existing filter logic)
    if filter_change_enabled:
        df = df.dropna(subset=[ratio_col])
        if filter_change_value is not None:
            try:
                filter_val = float(filter_change_value)
                df = df[df[ratio_col] > filter_val]
            except (ValueError, TypeError) as e:
                print(f"Invalid filter value for 'change >': {filter_change_value}. Error: {e}")
        if filter_change_lt_value is not None:
            try:
                filter_val = float(filter_change_lt_value)
                df = df[df[ratio_col] < filter_val]
            except (ValueError, TypeError) as e:
                print(f"Invalid filter value for 'change <': {filter_change_lt_value}. Error: {e}")
        if filter_change_from_value is not None and filter_change_to_value is not None:
            try:
                filter_from = float(filter_change_from_value)
                filter_to = float(filter_change_to_value)
                if filter_from <= filter_to:
                    df = df[(df[ratio_col] >= filter_from) & (df[ratio_col] <= filter_to)]
                else:
                    df = df[(df[ratio_col] >= filter_to) & (df[ratio_col] <= filter_from)]
            except (ValueError, TypeError) as e:
                print(f"Invalid filter values for 'change between': {filter_change_from_value}-{filter_change_to_value}. Error: {e}")
        elif filter_change_from_value is not None:
             try:
                filter_from = float(filter_change_from_value)
                df = df[df[ratio_col] >= filter_from]
             except (ValueError, TypeError) as e:
                print(f"Invalid filter value for 'change From': {filter_change_from_value}. Error: {e}")
        elif filter_change_to_value is not None:
             try:
                filter_to = float(filter_change_to_value)
                df = df[df[ratio_col] <= filter_to]
             except (ValueError, TypeError) as e:
                print(f"Invalid filter value for 'change To': {filter_change_to_value}. Error: {e}")

    # Apply comment filter if provided
    if filter_comment is not None and filter_comment.strip() != "":
        if 'comments' in df.columns:
            df = df[df['comments'].astype(str).str.lower().str.strip() == filter_comment.lower().strip()]
        else:
            df = df.head(0)

    # Get color status
    approved_cells = get_cell_color_status()

    # Apply ID filter if provided
    if filter_id is not None and filter_id != "":
        if number_col_exists:
            sample_type = type(df[number_col].iloc[0]) if not df.empty and not pd.isna(df[number_col].iloc[0]) else None
            if sample_type == int:
                try:
                    filter_id_value = int(filter_id)
                    df = df[df[number_col] == filter_id_value]
                except (ValueError, TypeError):
                    df = df[df[number_col].astype(str) == str(filter_id)]
            elif sample_type == float:
                try:
                    filter_id_value = float(filter_id)
                    df = df[df[number_col] == filter_id_value]
                except (ValueError, TypeError):
                    df = df[df[number_col].astype(str) == str(filter_id)]
            else:
                df = df[df[number_col].astype(str) == str(filter_id)]
        
        if len(df) == 0 or not number_col_exists:
            try:
                filter_idx = int(filter_id)
                if filter_idx in df.index:
                    df = df.loc[[filter_idx]]
            except (ValueError, TypeError):
                if not number_col_exists:
                    df = df.head(0)

    # Add color status info efficiently using vectorized operations
    df['col_a_approved'] = df.index.map(lambda idx: approved_cells.get(idx + 2, {}).get('col_a', False))
    df['col_a_type'] = df.index.map(lambda idx: approved_cells.get(idx + 2, {}).get('col_a_type', None))
    df['col_b_approved'] = df.index.map(lambda idx: approved_cells.get(idx + 2, {}).get('col_b', False))
    df['col_b_type'] = df.index.map(lambda idx: approved_cells.get(idx + 2, {}).get('col_b_type', None))

    # Apply Color Filters
    if filter_color_a != 'any':
        if filter_color_a == 'none':
            df = df[df['col_a_approved'] == False]
        else:
            df = df[(df['col_a_approved'] == True) & (df['col_a_type'] == filter_color_a)]

    if filter_color_b != 'any':
        if filter_color_b == 'none':
            df = df[df['col_b_approved'] == False]
        else:
            df = df[(df['col_b_approved'] == True) & (df['col_b_type'] == filter_color_b)]

    df = df.replace('_x000D_', '\n', regex=True).replace(r'\r\n|\r|\n', '\n', regex=True)

    total_rows = len(df)
    total_pages = math.ceil(total_rows / rows_per_page) if rows_per_page > 0 else 1
    page = max(1, min(page, total_pages))

    start_idx = (page - 1) * rows_per_page
    end_idx = start_idx + rows_per_page
    page_data = df.iloc[start_idx:end_idx]

    result = []
    original_indices = df.index[start_idx:end_idx]

    for i, df_idx in enumerate(original_indices):
        row = page_data.iloc[i]
        col_a, col_b = row[primary_text_col], row[secondary_text_col]

        # Handle case where values might be Series (e.g., from duplicate columns)
        if isinstance(col_a, pd.Series):
            col_a = col_a.iloc[0] if len(col_a) > 0 else None
        if isinstance(col_b, pd.Series):
            col_b = col_b.iloc[0] if len(col_b) > 0 else None

        row_id = row[number_col] if number_col_exists and number_col in row and pd.notna(row[number_col]) else df_idx

        if isinstance(col_a, str): 
            col_a = col_a.replace('_x000D_', '\n').replace('\r\n', '\n').replace('\r', '\n')
        if isinstance(col_b, str): 
            col_b = col_b.replace('_x000D_', '\n').replace('\r\n', '\n').replace('\r', '\n')

        highlighted_a, highlighted_b, status = compare_text(col_a, col_b)
        excel_row_idx = df_idx + 2
        row_approval = approved_cells.get(excel_row_idx, {'col_a': False, 'col_b': False, 'col_a_type': None, 'col_b_type': None})

        result.append({
            'row_idx': df_idx,
            'id': row_id,
            'col_a': col_a if not pd.isna(col_a) else "", 
            'col_b': col_b if not pd.isna(col_b) else "",
            'highlighted_a': highlighted_a, 
            'highlighted_b': highlighted_b, 
            'status': status,
            'col_a_approved': row_approval['col_a'], 
            'col_b_approved': row_approval['col_b'],
            'col_a_type': row_approval['col_a_type'], 
            'col_b_type': row_approval['col_b_type'],
            'ratio': row[ratio_col] if ratio_col in row else None
        })

    return result, total_pages, total_rows, change_col_exists

@app.route('/', methods=['GET'])
def index():
    global current_chunk

    # Get uploaded files list
    uploaded_files = get_uploaded_files_list()
    has_files = len(uploaded_files) > 0

    # Get current input file
    input_file = get_input_file_path()
    file_selected = input_file is not None

    # Determine the current file name for display
    current_file_name = None
    if file_selected:
        current_file_name = os.path.basename(input_file)

    # If no file is selected but files exist, show file selection view
    # If no files exist, show welcome/upload view
    # If file is selected, show data view

    # Default values for template
    data = []
    total_pages = 0
    total_rows = 0
    change_col_exists = False
    data_sheet_missing = False
    query_params = {}

    rows_per_page = request.args.get('rows_per_page', default=10, type=int)
    page = request.args.get('page', default=1, type=int)
    filter_change_enabled = request.args.get('filter_change_enabled') == 'on'
    filter_change_gt_value_str = request.args.get('filter_change_gt_value', default='').strip()
    filter_change_lt_value_str = request.args.get('filter_change_lt_value', default='').strip()
    filter_change_from_value_str = request.args.get('filter_change_from_value', default='').strip()
    filter_change_to_value_str = request.args.get('filter_change_to_value', default='').strip()
    filter_color_a = request.args.get('filter_color_a', default='any').strip().lower()
    filter_color_b = request.args.get('filter_color_b', default='any').strip().lower()
    sort_order = request.args.get('sort_order', default='asc').strip().lower()
    filter_id = request.args.get('filter_id', default=None)
    filter_comment = request.args.get('filter_comment', default=None)

    # If filter_id is provided but empty, set it to None
    if filter_id and filter_id.strip() == "":
        filter_id = None

    # If filter_comment is provided but empty, set it to None
    if filter_comment and filter_comment.strip() == "":
        filter_comment = None

    filter_change_gt_value = None
    filter_change_lt_value = None
    filter_change_from_value = None
    filter_change_to_value = None

    # Validate color filters
    valid_colors = ['any', 'none', 'green', 'red', 'yellow']
    if filter_color_a not in valid_colors: filter_color_a = 'any'
    if filter_color_b not in valid_colors: filter_color_b = 'any'

    # Validate sort_order
    valid_sort_orders = ['asc', 'desc', 'none']
    if sort_order not in valid_sort_orders: sort_order = 'asc'

    if filter_change_enabled:
        if filter_change_gt_value_str:
            try: filter_change_gt_value = float(filter_change_gt_value_str)
            except ValueError: print(f"Warning: Invalid 'change >' filter value: {filter_change_gt_value_str}")
        if filter_change_lt_value_str:
            try: filter_change_lt_value = float(filter_change_lt_value_str)
            except ValueError: print(f"Warning: Invalid 'change <' filter value: {filter_change_lt_value_str}")
        if filter_change_from_value_str:
            try: filter_change_from_value = float(filter_change_from_value_str)
            except ValueError: print(f"Warning: Invalid 'change From' filter value: {filter_change_from_value_str}")
        if filter_change_to_value_str:
            try: filter_change_to_value = float(filter_change_to_value_str)
            except ValueError: print(f"Warning: Invalid 'change To' filter value: {filter_change_to_value_str}")

        if filter_change_gt_value is None and filter_change_lt_value is None and filter_change_from_value is None and filter_change_to_value is None:
             filter_change_enabled = False

    # Only load data if a file is selected
    if file_selected and os.path.exists(input_file):
        try:
            wb = load_workbook(input_file, read_only=True)
            sheet_name = get_sheet_name()
            data_sheet_missing = sheet_name not in wb.sheetnames
            wb.close()
        except Exception: pass

        data, total_pages, total_rows, change_col_exists = get_excel_data(
            rows_per_page,
            page,
            filter_change_enabled,
            filter_change_gt_value,
            filter_change_lt_value,
            filter_change_from_value,
            filter_change_to_value,
            filter_color_a,
            filter_color_b,
            sort_order,
            filter_id,
            filter_comment
        )

        query_params = {
            'rows_per_page': rows_per_page,
        }
        if filter_change_enabled:
            query_params['filter_change_enabled'] = 'on'
            if filter_change_gt_value is not None: query_params['filter_change_value'] = filter_change_gt_value_str
            if filter_change_lt_value is not None: query_params['filter_change_lt_value'] = filter_change_lt_value_str
            if filter_change_from_value is not None: query_params['filter_change_from_value'] = filter_change_from_value_str
            if filter_change_to_value is not None: query_params['filter_change_to_value'] = filter_change_to_value_str

        # Add color filters to query params if they are not 'any'
        if filter_color_a != 'any': query_params['filter_color_a'] = filter_color_a
        if filter_color_b != 'any': query_params['filter_color_b'] = filter_color_b

        # Add sort_order to query params if it's not the default 'asc'
        if sort_order != 'asc': query_params['sort_order'] = sort_order

        # Add ID filter to query params if it's not None
        if filter_id is not None:
            query_params['filter_id'] = filter_id

        # Add comment filter to query params if it's not None
        if filter_comment is not None:
            query_params['filter_comment'] = filter_comment

    return render_template('index.html',
                          # State variables
                          has_files=has_files,
                          file_selected=file_selected,
                          uploaded_files=uploaded_files,
                          current_file_name=current_file_name,
                          # Data variables
                          data=data,
                          total_pages=total_pages,
                          current_page=page,
                          rows_per_page=rows_per_page,
                          total_rows=total_rows,
                          hadith_sheet_missing=data_sheet_missing,
                          filter_change_enabled=filter_change_enabled,
                          filter_change_gt_value=filter_change_gt_value_str,
                          filter_change_lt_value=filter_change_lt_value_str,
                          filter_change_from_value=filter_change_from_value_str,
                          filter_change_to_value=filter_change_to_value_str,
                          filter_color_a=filter_color_a,
                          filter_color_b=filter_color_b,
                          filter_id=filter_id,
                          filter_comment=filter_comment,
                          sort_order=sort_order,
                          change_col_exists=change_col_exists,
                          query_params=query_params,
                          available_chunks=get_available_chunks(),
                          current_chunk=current_chunk
                          )

@app.route('/edit', methods=['POST'])
def edit_cell():
    row_idx, new_text = request.form.get('row_idx', type=int), request.form.get('text', '')
    new_text = new_text.replace('<br>', '\n').replace('<br/>', '\n').replace('\r\n', '\n').replace('\r', '\n')

    input_file = get_input_file_path()
    if not input_file or not os.path.exists(input_file):
        return jsonify({'status': 'error', 'message': 'No file selected or file not found'})
    
    try:
        with file_lock:
            wb = safe_load_workbook(input_file)
            sheet_name = get_sheet_name()
            if sheet_name not in wb.sheetnames: 
                return jsonify({'status': 'error', 'message': f'{sheet_name} sheet not found in Excel file'})
            ws = wb[sheet_name]
            
            header = next(ws.rows)
            secondary_text_col_idx, primary_text_col_idx = 1, 0
            
            primary_text_col_name = get_column_name('primary_text')
            secondary_text_col_name = get_column_name('secondary_text')
            
            for idx, cell in enumerate(header):
                if cell.value == secondary_text_col_name:
                    secondary_text_col_idx = idx
                    break
            
            for idx, cell in enumerate(header):
                if cell.value == primary_text_col_name:
                    primary_text_col_idx = idx
                    break
            
            excel_row = row_idx + 2
            cell_address = f'{chr(65 + secondary_text_col_idx)}{excel_row}'
            ws[cell_address].value = new_text
            
            safe_save_workbook(wb, input_file)
            
            col_a_cell = ws[f'{chr(65 + primary_text_col_idx)}{excel_row}']
            col_a_text = str(col_a_cell.value) if col_a_cell.value is not None else ''
            highlighted_a, highlighted_b, status = compare_text(col_a_text, new_text)
            
            return jsonify({'status': 'success', 'highlighted_html': highlighted_b, 'diff_status': status})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/approve', methods=['POST'])
def approve_cell():
    row_idx = request.form.get('row_idx', type=int)
    column = request.form.get('column')
    approval_type = request.form.get('approval_type', 'green')
    
    input_file = get_input_file_path()
    try:
        if not input_file or not os.path.exists(input_file): 
            return jsonify({'status': 'error', 'message': 'Excel file not found'})
        
        with file_lock:
            wb = safe_load_workbook(input_file)
            sheet_name = get_sheet_name()
            if sheet_name not in wb.sheetnames: 
                return jsonify({'status': 'error', 'message': f'{sheet_name} sheet not found in Excel file'})
            
            ws = wb[sheet_name]
            header_row = next(ws.rows)
            
            primary_text_col_name = get_column_name('primary_text')
            secondary_text_col_name = get_column_name('secondary_text')
            
            primary_text_col_idx, secondary_text_col_idx = 0, 1
            
            for idx, cell in enumerate(header_row):
                col_name = cell.value
                if col_name == primary_text_col_name: 
                    primary_text_col_idx = idx
                elif col_name == secondary_text_col_name: 
                    secondary_text_col_idx = idx
            
            excel_row = row_idx + 2
            column_idx = primary_text_col_idx if column == 'a' else secondary_text_col_idx
            cell_address = f'{chr(65 + column_idx)}{excel_row}'
            
            colors = {'green': "00FF00", 'yellow': "FFFF00", 'red': "FFFF0000"}
            color = colors.get(approval_type, "00FF00")
            
            ws[cell_address].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            safe_save_workbook(wb, input_file)
            
            return jsonify({'status': 'success', 'message': 'Cell approved successfully', 
                           'row_idx': row_idx, 'column': column, 'approval_type': approval_type})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/reset_cell', methods=['POST'])
def reset_cell():
    row_idx, column = request.form.get('row_idx', type=int), request.form.get('column')
    
    input_file = get_input_file_path() # Get path from config
    if not input_file or not os.path.exists(input_file): return jsonify({'status': 'error', 'message': 'Excel file not found'})
    
    try:
        wb = load_workbook(input_file)
        sheet_name = get_sheet_name()
        if sheet_name not in wb.sheetnames: return jsonify({'status': 'error', 'message': f'{sheet_name} sheet not found in Excel file'})
        
        ws = wb[sheet_name]
        header_row = next(ws.rows)
        
        primary_text_col_name = get_column_name('primary_text')
        secondary_text_col_name = get_column_name('secondary_text')
        
        primary_text_col_idx, secondary_text_col_idx = 0, 1
        
        for idx, cell in enumerate(header_row):
            col_name = cell.value
            if col_name == primary_text_col_name: primary_text_col_idx = idx
            elif col_name == secondary_text_col_name: secondary_text_col_idx = idx
        
        excel_row = row_idx + 2
        column_idx = primary_text_col_idx if column == 'a' else secondary_text_col_idx
        cell_address = f'{chr(65 + column_idx)}{excel_row}'
        
        ws[cell_address].fill = PatternFill(fill_type=None)
        wb.save(input_file)
        
        return jsonify({'status': 'success', 'message': 'Cell color reset successfully', 'row_idx': row_idx, 'column': column})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/save_selection', methods=['POST'])
def save_selection():
    selected_text = request.form.get('selected_text', '')
    row_idx = request.form.get('row_idx', type=int)
    
    if not selected_text or row_idx is None:
        return jsonify({'status': 'error', 'message': 'Missing required data'})
    
    try:
        # Load existing selections or create new
        selections_file = 'selections.xlsx'
        
        if os.path.exists(selections_file):
            df = pd.read_excel(selections_file)
        else:
            df = pd.DataFrame(columns=['row_idx', 'selected_text', 'timestamp'])
        
        # Add new selection
        new_selection = pd.DataFrame({
            'row_idx': [row_idx],
            'selected_text': [selected_text],
            'timestamp': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        })
        
        df = pd.concat([df, new_selection], ignore_index=True)
        
        # Save to Excel
        df.to_excel(selections_file, index=False)
        
        return jsonify({'status': 'success', 'message': 'Selection saved successfully'})
    
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Error saving selection: {str(e)}'})

@app.route('/keep_this', methods=['POST'])
def keep_this():
    row_idx = request.form.get('row_idx', type=int)
    diff_id = request.form.get('diff_id', '')
    
    if row_idx is None or not diff_id:
        return jsonify({'status': 'error', 'message': 'Missing required data'})
    
    try:
        input_file = get_input_file_path()
        if not input_file or not os.path.exists(input_file):
            return jsonify({'status': 'error', 'message': 'Excel file not found'})
        
        with file_lock:
            wb = safe_load_workbook(input_file)
            sheet_name = get_sheet_name()
            if sheet_name not in wb.sheetnames:
                return jsonify({'status': 'error', 'message': f'{sheet_name} sheet not found in Excel file'})
            ws = wb[sheet_name]
            
            # Get column indices
            header = next(ws.rows)
            primary_text_col_idx, secondary_text_col_idx = 0, 1
            
            primary_text_col_name = get_column_name('primary_text')
            secondary_text_col_name = get_column_name('secondary_text')
            
            for idx, cell in enumerate(header):
                if cell.value == primary_text_col_name:
                    primary_text_col_idx = idx
                elif cell.value == secondary_text_col_name:
                    secondary_text_col_idx = idx
            
            excel_row = row_idx + 2
            
            # Get current texts
            col_a_cell = ws[f'{get_column_letter(primary_text_col_idx + 1)}{excel_row}']
            col_b_cell = ws[f'{get_column_letter(secondary_text_col_idx + 1)}{excel_row}']
            
            col_a_text = str(col_a_cell.value) if col_a_cell.value is not None else ''
            col_b_text = str(col_b_cell.value) if col_b_cell.value is not None else ''
            
            # Perform the selective replacement
            new_col_b_text = perform_selective_replacement(col_a_text, col_b_text, diff_id)
            
            # Update Column B in Excel
            col_b_cell.value = new_col_b_text
            
            # Clear any existing fill color for Column B
            col_b_cell.fill = PatternFill()
            
            safe_save_workbook(wb, input_file)
            
            # Generate new comparison for response
            highlighted_a, highlighted_b, status = compare_text(col_a_text, new_col_b_text)
            
            # Get color status
            color_status = get_cell_color_status()
            row_approval = color_status.get(excel_row, {'col_b': False, 'col_b_type': None})
            col_b_approved = row_approval['col_b']
            col_b_type = row_approval['col_b_type']
            
            return jsonify({
                'status': 'success',
                'new_text': new_col_b_text,
                'new_content': highlighted_b,
                'highlighted_html': highlighted_b,
                'highlighted_a_html': highlighted_a,
                'diff_status': status,
                'col_b_approved': col_b_approved,
                'col_b_type': col_b_type
            })
        
    except Exception as e:
        print(f"Error in keep_this for row {row_idx}: {type(e).__name__} - {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'An unexpected error occurred: {str(e)}'})

def perform_selective_replacement(col_a_text, col_b_text, target_diff_id):
    """
    Perform selective replacement of a specific diff in Column B with content from Column A
    """
    if pd.isna(col_a_text) or pd.isna(col_b_text):
        return col_b_text
    
    col_a_text, col_b_text = str(col_a_text), str(col_b_text)
    if col_a_text == col_b_text:
        return col_b_text
    
    # Use the same preprocessing as in compare_text
    line_break_marker = " ¶ "
    col_a_prep = col_a_text.replace('\r\n', '\n').replace('\r', '\n').replace('\n', line_break_marker)
    col_b_prep = col_b_text.replace('\r\n', '\n').replace('\r', '\n').replace('\n', line_break_marker)
    
    words_a = re.split(r'(\s+)', col_a_prep)
    words_b = re.split(r'(\s+)', col_b_prep)
    words_a = [word for word in words_a if word]
    words_b = [word for word in words_b if word]
    
    matcher = difflib.SequenceMatcher(None, words_a, words_b, autojunk=False)
    
    # Build the result by processing opcodes
    result_words = []
    diff_id_counter = 0
    
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        words_a_segment = "".join(words_a[i1:i2])
        words_b_segment = "".join(words_b[j1:j2])
        
        if tag == 'equal':
            result_words.append(words_b_segment)
        else:
            # Generate the same deterministic diff_id as in compare_text
            diff_id = f"diff-{diff_id_counter}-{tag}-{i1}-{i2}-{j1}-{j2}"
            diff_id_counter += 1
            
            # If this is the target diff_id we want to replace
            if diff_id == target_diff_id:
                # Use content from Column A instead of Column B
                result_words.append(words_a_segment)
            else:
                # Keep the original Column B content
                result_words.append(words_b_segment)
    
    # Convert back to original format
    final_text = "".join(result_words).replace(line_break_marker.strip(), "\n")
    return final_text

@app.route('/preview_diff', methods=['POST'])
def preview_diff():
    text1, text2 = request.form.get('text1', ''), request.form.get('text2', '')
    try:
        highlighted_a, highlighted_b, status = compare_text(text1, text2)
        return jsonify({'highlighted_a': highlighted_a, 'highlighted_b': highlighted_b, 'status': status})
    except Exception as e:
        return jsonify({'status': 'error', 'message': f"Error comparing texts: {str(e)}"})

@app.route('/regenerate_cell', methods=['POST'])
def regenerate_cell():
    row_idx = request.form.get('row_idx', type=int)
    provider = request.form.get('provider', 'google')
    try:
        new_text = generate(row_idx, get_input_file_path(), provider=provider).strip()

        input_file = get_input_file_path()
        
        if not input_file or not os.path.exists(input_file):
            return jsonify({'status': 'error', 'message': 'Excel file not found after regeneration'})

        with file_lock:
            wb = safe_load_workbook(input_file)
            sheet_name = get_sheet_name()
            if sheet_name not in wb.sheetnames:
                return jsonify({'status': 'error', 'message': f'{sheet_name} sheet not found in Excel file'})
            ws = wb[sheet_name]

            header = next(ws.rows)
            primary_text_col_name = get_column_name('primary_text')
            secondary_text_col_name = get_column_name('secondary_text')
            
            secondary_text_col_idx = 1
            primary_text_col_idx = 0
            
            for idx, cell in enumerate(header):
                if cell.value == secondary_text_col_name:
                    secondary_text_col_idx = idx
                elif cell.value == primary_text_col_name:
                    primary_text_col_idx = idx

            excel_row = row_idx + 2
            cell_address = f'{get_column_letter(secondary_text_col_idx + 1)}{excel_row}'
            ws[cell_address].value = new_text
            ws[cell_address].fill = PatternFill(fill_type=None)  # Clear existing fill

            safe_save_workbook(wb, input_file)

            # Fetch updated color status for Column B
            color_status = get_cell_color_status()
            row_approval = color_status.get(excel_row, {'col_b': False, 'col_b_type': None})
            col_b_approved = row_approval['col_b']
            col_b_type = row_approval['col_b_type']

            # Get original text from Column A for comparison
            col_a_cell = ws[f'{get_column_letter(primary_text_col_idx + 1)}{excel_row}']
            col_a_text = str(col_a_cell.value) if col_a_cell.value is not None else ''
            highlighted_a, highlighted_b, status = compare_text(col_a_text, new_text)

            return jsonify({
                'status': 'success',
                'new_text': new_text,
                'highlighted_html': highlighted_b,
                'highlighted_a_html': highlighted_a,  # Include highlighted HTML for column A
                'diff_status': status,
                'col_b_approved': col_b_approved,  # Add approval status
                'col_b_type': col_b_type          # Add approval type
            })

    except FileNotFoundError as e:
        return jsonify({'status': 'error', 'message': str(e)})
    except ValueError as e:
        return jsonify({'status': 'error', 'message': str(e)})
    except Exception as e:
        print(f"Error during regeneration or file update for row {row_idx}: {type(e).__name__} - {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'An unexpected error occurred: {str(e)}'})

@app.route('/regenerate_multiple_cells', methods=['POST'])
def regenerate_multiple_cells():
    row_ids = request.json.get('row_ids', [])
    provider = request.json.get('provider', 'google')

    print(f"Regenerating rows: {row_ids}")
    
    if not row_ids:
        return jsonify({'status': 'error', 'message': 'No row IDs provided'})
    
    try:
        import concurrent.futures
        
        input_file = get_input_file_path()
        results = []
        
        if not input_file or not os.path.exists(input_file):
            return jsonify({'status': 'error', 'message': 'Excel file not found'})
        
        # First, generate all the new texts in parallel
        generated_texts = {}
        
        def generate_text_for_row(row_idx):
            try:
                print(f"Generating text for row: {row_idx}")
                new_text = generate(row_idx, input_file, provider=provider).strip()
                return {'status': 'success', 'row_idx': row_idx, 'new_text': new_text}
            except Exception as e:
                import traceback
                return {
                    'status': 'error',
                    'row_idx': row_idx,
                    'message': str(e),
                    'traceback': traceback.format_exc()
                }
        
        # Generate all texts in parallel
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(10, len(row_ids))) as executor:
            # Submit all generation tasks
            future_to_row = {executor.submit(generate_text_for_row, row_idx): row_idx for row_idx in row_ids}
            
            # Collect results as they complete
            for future in concurrent.futures.as_completed(future_to_row):
                row_idx = future_to_row[future]
                try:
                    result = future.result()
                    if result['status'] == 'success':
                        generated_texts[row_idx] = result['new_text']
                    else:
                        results.append(result)  # Store error results
                except Exception as e:
                    results.append({
                        'status': 'error',
                        'row_idx': row_idx,
                        'message': str(e)
                    })
        
        # If there are successful generations, update the Excel file only once
        if generated_texts:
            try:
                # Use the new batch update function
                batch_results = batch_update_excel_cells(input_file, generated_texts)
                results.extend(batch_results)
                
            except Exception as e:
                import traceback
                error_message = f"Error updating Excel file: {str(e)}"
                traceback.print_exc()
                
                # Add error for each row that was not already recorded as an error
                for row_idx in generated_texts.keys():
                    if not any(r.get('row_idx') == row_idx and r.get('status') == 'error' for r in results):
                        results.append({
                            'status': 'error',
                            'row_idx': row_idx,
                            'message': error_message
                        })
        
        success_count = sum(1 for r in results if r['status'] == 'success')
        
        return jsonify({
            'status': 'success',
            'message': f'Successfully processed {success_count} of {len(row_ids)} rows',
            'results': results
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': f'An unexpected error occurred: {str(e)}'
        })

@app.route('/get_comment', methods=['GET'])
def get_comment():
    row_idx = request.args.get('row_idx', type=int)
    
    input_file = get_input_file_path() # Get path from config
    if not input_file or not os.path.exists(input_file): # Use configured path
        return jsonify({'status': 'error', 'message': 'Excel file not found'})
    
    try:
        sheet_name = get_sheet_name()
        df = pd.read_excel(input_file, sheet_name=sheet_name)
        
        if 'comments' not in df.columns:
            return jsonify({'comment': '', 'status': 'success'})
        
        comment = df.loc[row_idx, 'comments'] if pd.notna(df.loc[row_idx, 'comments']) else ''
        
        return jsonify({'comment': comment, 'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/save_comment', methods=['POST'])
def save_comment():
    row_idx = request.form.get('row_idx', type=int)
    comment = request.form.get('comment', '')
    
    input_file = get_input_file_path() # Get path from config
    if not input_file or not os.path.exists(input_file): # Use configured path
        return jsonify({'status': 'error', 'message': 'Excel file not found'})
    
    try:
        wb = load_workbook(input_file)
        sheet_name = get_sheet_name()
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        
        header_row = next(ws.rows)
        comments_col_idx = None
        
        for idx, cell in enumerate(header_row):
            if cell.value == 'comments':
                comments_col_idx = idx
                break
        
        if comments_col_idx is None:
            comments_col_idx = len(header_row)
            comments_col_letter = get_column_letter(comments_col_idx + 1)
            ws[f'{comments_col_letter}1'] = 'comments'
        
        excel_row = row_idx + 2
        comments_col_letter = get_column_letter(comments_col_idx + 1)
        ws[f'{comments_col_letter}{excel_row}'] = comment
        
        wb.save(input_file)
        
        return jsonify({'status': 'success', 'message': 'Comment saved successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})
    

@app.route('/get_arabic_text', methods=['GET'])
def get_arabic_text():
    try:
        row_idx = request.args.get('row_idx')
        
        if not row_idx:
            return jsonify({'status': 'error', 'message': 'Row index is required'})
        
        input_file = get_input_file_path()
        if not input_file or not os.path.exists(input_file):
            return jsonify({'status': 'error', 'message': 'Input file not found'})
        
        try:
            # Read the Excel file
            xls = pd.ExcelFile(input_file, engine='openpyxl')
            sheet_name = get_sheet_name()
            sheet_name = sheet_name if sheet_name in xls.sheet_names else xls.sheet_names[0]
            df = pd.read_excel(xls, sheet_name=sheet_name)
            xls.close()  # Close the file handle to prevent WinError 32
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Error reading Excel file: {str(e)}'})
        
        # Get the Arabic column name from config
        arabic_column = get_column_name('arabic_text')
        
        # Check for Arabic column in order of preference
        if arabic_column not in df.columns:
            # First check for common Arabic column names if config value not found
            if 'arabic_text' in df.columns:
                arabic_column = 'arabic_text'
            elif 'hadith_arabic' in df.columns:
                arabic_column = 'hadith_arabic'
            else:
                # Fallback to any column with 'arabic' in the name
                for col in df.columns:
                    if 'arabic' in str(col).lower():
                        arabic_column = col
                        break
        
        # If still no Arabic column found, return an error
        if arabic_column is None or arabic_column not in df.columns:
            return jsonify({
                'status': 'error', 
                'message': 'No Arabic text column found. Available columns: ' + ', '.join(df.columns)
            })
        
        # Convert row_idx to integer
        try:
            row_idx = int(row_idx)
            
            # Directly use the row index as provided (excel row number - 2)
            df_row_idx = row_idx - 2
            
            print(f"Requested row: {row_idx}, DataFrame index: {df_row_idx}, Max rows: {len(df)}")
            
            if df_row_idx < 0 or df_row_idx >= len(df):
                return jsonify({
                    'status': 'error', 
                    'message': f'Row index {row_idx} out of range (should be between 2 and {len(df)+1})'
                })
            
            # Get the Arabic text from the row
            arabic_text = df.iloc[df_row_idx][arabic_column]
            
            # Handle NaN or None values
            if pd.isna(arabic_text):
                arabic_text = "لا يوجد نص عربي" # "No Arabic text available" in Arabic
            else:
                arabic_text = str(arabic_text)
            
            return jsonify({
                'status': 'success',
                'arabic_text': arabic_text,
                'row_used': row_idx
            })
            
        except ValueError:
            return jsonify({'status': 'error', 'message': f'Invalid row index: {row_idx}'})
        
    except Exception as e:
        import traceback
        print(f"Error retrieving Arabic text: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/translate_arabic_to_bangla', methods=['GET'])
def translate_arabic_to_bangla():
    try:
        row_idx = request.args.get('row_idx', type=int)
        provider = request.args.get('provider', 'google')
        
        if row_idx is None:
            return jsonify({'status': 'error', 'message': 'Row index is required'})

        input_file = get_input_file_path()
        if not input_file or not os.path.exists(input_file):
            return jsonify({'status': 'error', 'message': 'Input file not found'})

        # Read the Excel file
        xls = pd.ExcelFile(input_file, engine='openpyxl')
        sheet_name = get_sheet_name()
        sheet_name = sheet_name if sheet_name in xls.sheet_names else xls.sheet_names[0]
        df = pd.read_excel(xls, sheet_name=sheet_name)
        xls.close()  # Close the file handle to prevent WinError 32

        # Get the Arabic column name from config
        arabic_column = get_column_name('arabic_text')
        
        # Check for Arabic column
        if arabic_column not in df.columns:
            if 'arabic_text' in df.columns:
                arabic_column = 'arabic_text'
            elif 'hadith_arabic' in df.columns:
                arabic_column = 'hadith_arabic'
            else:
                for col in df.columns:
                    if 'arabic' in str(col).lower():
                        arabic_column = col
                        break
        
        if arabic_column is None or arabic_column not in df.columns:
            return jsonify({
                'status': 'error',
                'message': 'No Arabic text column found. Available columns: ' + ', '.join(df.columns)
            })

        # Validate row index
        df_row_idx = row_idx - 2  # Adjust for 0-based indexing and header row
        if df_row_idx < 0 or df_row_idx >= len(df):
            return jsonify({
                'status': 'error',
                'message': f'Row index {row_idx} out of range (should be between 2 and {len(df)+1})'
            })

        # Get the Arabic text
        arabic_text = df.iloc[df_row_idx][arabic_column]
        arabic_text = str(arabic_text) if not pd.isna(arabic_text) else "لا يوجد نص عربي"

        # Prepare the translation query
        from src.prompt import inject_variables, translate_arabic_to_bangla_prompt
        query = inject_variables(translate_arabic_to_bangla_prompt, {
            "arabic_text": arabic_text
        })

        # Call the AI model for translation using specified provider
       
        translated_text = ask(query, provider=provider).text.strip()

        return jsonify({
            'status': 'success',
            'arabic_text': arabic_text,
            'translated_bangla': translated_text,
            'row_used': row_idx
        })

    except Exception as e:
        import traceback
        print(f"Error translating Arabic to Bangla: {str(e)}")
        print(traceback.format_exc())
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/recalculate_ratios', methods=['POST'])
def recalculate_ratios():
    input_file = get_input_file_path()
    if not input_file or not os.path.exists(input_file):
        return jsonify({'status': 'error', 'message': 'Excel file not found'})
        
    try:
        # Get sheet and column names from config
        sheet_name = get_sheet_name()
        primary_text_col = get_column_name('primary_text')
        secondary_text_col = get_column_name('secondary_text')
        ratio_col = get_column_name('ratio')
        
        # Read the Excel file
        df = pd.read_excel(input_file, sheet_name=sheet_name)
        
        # Calculate ratios for each row
        df[ratio_col] = df.apply(lambda row: difflib.SequenceMatcher(
            None, 
            extract_standard_letters(str(row[primary_text_col]) if pd.notna(row[primary_text_col]) else ""),
            extract_standard_letters(str(row[secondary_text_col]) if pd.notna(row[secondary_text_col]) else ""),
            autojunk=False
        ).ratio() * 100, axis=1)
        
        # Save the updated ratios back to Excel
        wb = load_workbook(input_file)
        if sheet_name not in wb.sheetnames:
            return jsonify({'status': 'error', 'message': f"'{sheet_name}' sheet not found in Excel file"})
            
        ws = wb[sheet_name]
        
        # Find the last column index or the existing ratio column
        ratio_col_idx = None
        last_col_idx = len(next(ws.rows))
        
        for idx, cell in enumerate(next(ws.rows)):
            if cell.value == ratio_col:
                ratio_col_idx = idx
                break
                
        ratio_col_letter = get_column_letter(ratio_col_idx + 1) if ratio_col_idx is not None else get_column_letter(last_col_idx + 1)
        
        # Add or update ratio header if needed
        if ratio_col_idx is None:
            ws[f'{ratio_col_letter}1'] = ratio_col
        
        # Update ratio values for each row
        for idx, ratio in enumerate(df[ratio_col], start=2):
            ws[f'{ratio_col_letter}{idx}'] = ratio
        
        wb.save(input_file)
        
        return jsonify({'status': 'success', 'message': 'Ratios recalculated successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/select_chunk', methods=['POST'])
def select_chunk():
    global current_chunk
    chunk_path = request.form.get('chunk_path')
    
    # Validate the chunk path
    if chunk_path and os.path.exists(chunk_path) and os.path.isfile(chunk_path):
        current_chunk = chunk_path
        print(f"Selected chunk changed to: {current_chunk}")
    else:
        print(f"Warning: Invalid chunk path: {chunk_path}")
    
    # Redirect back to the index page with the same query parameters
    return redirect(url_for('index', **request.args))

@app.context_processor
def utility_processor():
    return dict(urlencode=urlencode)

@app.route('/regenerate_with_prompt_1', methods=['POST'])
def regenerate_with_prompt_1():
    row_idx = request.form.get('row_idx', type=int)
    provider = request.form.get('provider', 'google')
    try:
        input_file = get_input_file_path()
        
        # 1. Read row data
        from src.generate_cell import read_row
        arabic_text, _, current_bangla = read_row(row_idx, input_file)
        
        # 2. Generate text using prompt 1
        from src.prompt import inject_variables, read_file
        from src.ai import ask
        query = inject_variables(read_file("./prompts/1.md"), {
            "hadis_arabic_text": arabic_text,
            "previous_generated_bangla": current_bangla
        })
        new_text = ask(query, provider=provider).text.strip()

        # 3. Update Excel file with new text
        if not input_file or not os.path.exists(input_file):
            return jsonify({'status': 'error', 'message': 'Excel file not found after regeneration'})

        wb = load_workbook(input_file)
        sheet_name = get_sheet_name()
        if sheet_name not in wb.sheetnames:
            return jsonify({'status': 'error', 'message': f'{sheet_name} sheet not found in Excel file'})
        ws = wb[sheet_name]

        header = next(ws.rows)
        primary_text_col_name = get_column_name('primary_text')
        secondary_text_col_name = get_column_name('secondary_text')
        
        secondary_text_col_idx = 1
        primary_text_col_idx = 0
        
        for idx, cell in enumerate(header):
            if cell.value == secondary_text_col_name:
                secondary_text_col_idx = idx
            elif cell.value == primary_text_col_name:
                primary_text_col_idx = idx

        excel_row = row_idx + 2
        cell_address = f'{get_column_letter(secondary_text_col_idx + 1)}{excel_row}'
        ws[cell_address].value = new_text
        ws[cell_address].fill = PatternFill(fill_type=None)  # Clear existing fill

        wb.save(input_file)

        # 4. Get updated color status and prepare comparison data
        color_status = get_cell_color_status()
        row_approval = color_status.get(excel_row, {'col_b': False, 'col_b_type': None})
        col_b_approved = row_approval['col_b']
        col_b_type = row_approval['col_b_type']

        # Get original text from Column A for comparison
        col_a_cell = ws[f'{get_column_letter(primary_text_col_idx + 1)}{excel_row}']
        col_a_text = str(col_a_cell.value) if col_a_cell.value is not None else ''
        highlighted_a, highlighted_b, status = compare_text(col_a_text, new_text)

        return jsonify({
            'status': 'success',
            'new_text': new_text,
            'highlighted_html': highlighted_b,
            'highlighted_a_html': highlighted_a,  # Include highlighted HTML for column A
            'diff_status': status,
            'col_b_approved': col_b_approved,  # Add approval status
            'col_b_type': col_b_type          # Add approval type
        })

    except FileNotFoundError as e:
        return jsonify({'status': 'error', 'message': str(e)})
    except ValueError as e:
        return jsonify({'status': 'error', 'message': str(e)})
    except Exception as e:
        print(f"Error during regeneration or file update for row {row_idx}: {type(e).__name__} - {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'An unexpected error occurred: {str(e)}'})


@app.route('/regenerate_with_prompt_2', methods=['POST'])
def regenerate_with_prompt_2():
    row_idx = request.form.get('row_idx', type=int)
    provider = request.form.get('provider', 'google')
    try:
        input_file = get_input_file_path()
        
        # 1. Read row data
        from src.generate_cell import read_row
        arabic_text, _, current_bangla = read_row(row_idx, input_file)
        
        # 2. Generate text using prompt 2
        from src.prompt import inject_variables, read_file
        from src.ai import ask
        query = inject_variables(read_file("./prompts/2.md"), {
            "hadis_arabic_text": arabic_text,
            "previous_generated_bangla": current_bangla
        })
        new_text = ask(query, provider=provider).text.strip()

        # 3. Update Excel file with new text
        if not input_file or not os.path.exists(input_file):
            return jsonify({'status': 'error', 'message': 'Excel file not found after regeneration'})

        wb = load_workbook(input_file)
        sheet_name = get_sheet_name()
        if sheet_name not in wb.sheetnames:
            return jsonify({'status': 'error', 'message': f'{sheet_name} sheet not found in Excel file'})
        ws = wb[sheet_name]

        header = next(ws.rows)
        primary_text_col_name = get_column_name('primary_text')
        secondary_text_col_name = get_column_name('secondary_text')
        
        secondary_text_col_idx = 1
        primary_text_col_idx = 0
        
        for idx, cell in enumerate(header):
            if cell.value == secondary_text_col_name:
                secondary_text_col_idx = idx
            elif cell.value == primary_text_col_name:
                primary_text_col_idx = idx

        excel_row = row_idx + 2
        cell_address = f'{get_column_letter(secondary_text_col_idx + 1)}{excel_row}'
        ws[cell_address].value = new_text
        ws[cell_address].fill = PatternFill(fill_type=None)  # Clear existing fill

        wb.save(input_file)

        # 4. Get updated color status and prepare comparison data
        color_status = get_cell_color_status()
        row_approval = color_status.get(excel_row, {'col_b': False, 'col_b_type': None})
        col_b_approved = row_approval['col_b']
        col_b_type = row_approval['col_b_type']

        # Get original text from Column A for comparison
        col_a_cell = ws[f'{get_column_letter(primary_text_col_idx + 1)}{excel_row}']
        col_a_text = str(col_a_cell.value) if col_a_cell.value is not None else ''
        highlighted_a, highlighted_b, status = compare_text(col_a_text, new_text)

        return jsonify({
            'status': 'success',
            'new_text': new_text,
            'highlighted_html': highlighted_b,
            'highlighted_a_html': highlighted_a,  # Include highlighted HTML for column A
            'diff_status': status,
            'col_b_approved': col_b_approved,  # Add approval status
            'col_b_type': col_b_type          # Add approval type
        })

    except FileNotFoundError as e:
        return jsonify({'status': 'error', 'message': str(e)})
    except ValueError as e:
        return jsonify({'status': 'error', 'message': str(e)})
    except Exception as e:
        print(f"Error during regeneration or file update for row {row_idx}: {type(e).__name__} - {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'An unexpected error occurred: {str(e)}'})


@app.route('/regenerate_multiple_with_prompt_1', methods=['POST'])
def regenerate_multiple_with_prompt_1():
    row_ids = request.json.get('row_ids', [])
    provider = request.json.get('provider', 'google')

    print(f"Regenerating rows with prompt 1: {row_ids}")
    
    if not row_ids:
        return jsonify({'status': 'error', 'message': 'No row IDs provided'})
    
    try:
        import concurrent.futures
        
        input_file = get_input_file_path()
        results = []
        
        if not input_file or not os.path.exists(input_file):
            return jsonify({'status': 'error', 'message': 'Excel file not found'})
        
        # First, generate all the new texts in parallel
        generated_texts = {}
        
        def generate_text_for_row_prompt_1(row_idx):
            try:
                print(f"Generating text with prompt 1 for row: {row_idx}")
                
                # Read row data
                from src.generate_cell import read_row
                arabic_text, _, current_bangla = read_row(row_idx, input_file)
                
                # Generate text using prompt 1
                from src.prompt import inject_variables, read_file
                from src.ai import ask
                query = inject_variables(read_file("./prompts/1.md"), {
                    "hadis_arabic_text": arabic_text,
                    "previous_generated_bangla": current_bangla
                })
                new_text = ask(query, provider=provider).text.strip()
                
                return {'status': 'success', 'row_idx': row_idx, 'new_text': new_text}
            except Exception as e:
                import traceback
                return {
                    'status': 'error',
                    'row_idx': row_idx,
                    'message': str(e),
                    'traceback': traceback.format_exc()
                }
        
        # Generate all texts in parallel
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(10, len(row_ids))) as executor:
            # Submit all generation tasks
            future_to_row = {executor.submit(generate_text_for_row_prompt_1, row_idx): row_idx for row_idx in row_ids}
            
            # Collect results as they complete
            for future in concurrent.futures.as_completed(future_to_row):
                row_idx = future_to_row[future]
                try:
                    result = future.result()
                    if result['status'] == 'success':
                        generated_texts[row_idx] = result['new_text']
                    else:
                        results.append(result)  # Store error results
                except Exception as e:
                    results.append({
                        'status': 'error',
                        'row_idx': row_idx,
                        'message': str(e)
                    })
        
        # If there are successful generations, update the Excel file only once
        if generated_texts:
            try:
                # Use the new batch update function
                batch_results = batch_update_excel_cells(input_file, generated_texts)
                results.extend(batch_results)
                
            except Exception as e:
                import traceback
                error_message = f"Error updating Excel file: {str(e)}"
                traceback.print_exc()
                
                # Add error for each row that was not already recorded as an error
                for row_idx in generated_texts.keys():
                    if not any(r.get('row_idx') == row_idx and r.get('status') == 'error' for r in results):
                        results.append({
                            'status': 'error',
                            'row_idx': row_idx,
                            'message': error_message
                        })
        
        # Return all results
        return jsonify({
            'status': 'success',
            'message': f'Completed regeneration with prompt 1 for {len(generated_texts)} rows. {len(row_ids) - len(generated_texts)} failed.',
            'results': results
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': f'An unexpected error occurred: {str(e)}'
        })


@app.route('/regenerate_multiple_with_prompt_2', methods=['POST'])
def regenerate_multiple_with_prompt_2():
    row_ids = request.json.get('row_ids', [])
    provider = request.json.get('provider', 'google')

    print(f"Regenerating rows with prompt 2: {row_ids}")
    
    if not row_ids:
        return jsonify({'status': 'error', 'message': 'No row IDs provided'})
    
    try:
        import concurrent.futures
        
        input_file = get_input_file_path()
        results = []
        
        if not input_file or not os.path.exists(input_file):
            return jsonify({'status': 'error', 'message': 'Excel file not found'})
        
        # First, generate all the new texts in parallel
        generated_texts = {}
        
        def generate_text_for_row_prompt_2(row_idx):
            try:
                print(f"Generating text with prompt 2 for row: {row_idx}")
                
                # Read row data
                from src.generate_cell import read_row
                arabic_text, _, current_bangla = read_row(row_idx, input_file)
                
                # Generate text using prompt 2
                from src.prompt import inject_variables, read_file
                from src.ai import ask
                query = inject_variables(read_file("./prompts/2.md"), {
                    "hadis_arabic_text": arabic_text,
                    "previous_generated_bangla": current_bangla
                })
                new_text = ask(query, provider=provider).text.strip()
                
                return {'status': 'success', 'row_idx': row_idx, 'new_text': new_text}
            except Exception as e:
                import traceback
                return {
                    'status': 'error',
                    'row_idx': row_idx,
                    'message': str(e),
                    'traceback': traceback.format_exc()
                }
        
        # Generate all texts in parallel
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(10, len(row_ids))) as executor:
            # Submit all generation tasks
            future_to_row = {executor.submit(generate_text_for_row_prompt_2, row_idx): row_idx for row_idx in row_ids}
            
            # Collect results as they complete
            for future in concurrent.futures.as_completed(future_to_row):
                row_idx = future_to_row[future]
                try:
                    result = future.result()
                    if result['status'] == 'success':
                        generated_texts[row_idx] = result['new_text']
                    else:
                        results.append(result)  # Store error results
                except Exception as e:
                    results.append({
                        'status': 'error',
                        'row_idx': row_idx,
                        'message': str(e)
                    })
        
        # If there are successful generations, update the Excel file only once
        if generated_texts:
            try:
                # Use the new batch update function
                batch_results = batch_update_excel_cells(input_file, generated_texts)
                results.extend(batch_results)
                
            except Exception as e:
                import traceback
                error_message = f"Error updating Excel file: {str(e)}"
                traceback.print_exc()
                
                # Add error for each row that was not already recorded as an error
                for row_idx in generated_texts.keys():
                    if not any(r.get('row_idx') == row_idx and r.get('status') == 'error' for r in results):
                        results.append({
                            'status': 'error',
                            'row_idx': row_idx,
                            'message': error_message
                        })
        
        # Return all results
        return jsonify({
            'status': 'success',
            'message': f'Completed regeneration with prompt 2 for {len(generated_texts)} rows. {len(row_ids) - len(generated_texts)} failed.',
            'results': results
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': f'An unexpected error occurred: {str(e)}'
        })

@app.route('/regenerate_with_custom_prompt', methods=['POST'])
def regenerate_with_custom_prompt():
    row_idx = request.form.get('row_idx', type=int)
    custom_prompt = request.form.get('prompt', '')
    provider = request.form.get('provider', 'google')
    
    if not custom_prompt.strip():
        return jsonify({'status': 'error', 'message': 'Empty prompt provided'})
    
    try:
        input_file = get_input_file_path()
        
        # 1. Read row data
        from src.generate_cell import read_row
        arabic_text, col_a_text, col_b_text = read_row(row_idx, input_file)
        
        # 2. Process the custom prompt by replacing placeholders
        processed_prompt = inject_variables(custom_prompt, {
            "arabic_text": arabic_text,
            "col_a_text": col_a_text,
            "col_b_text": col_b_text
        })
        
        # 3. Generate text using the custom prompt and selected provider
        from src.ai import ask
        new_text = ask(processed_prompt, provider=provider).text.strip()

        # 4. Update Excel file with new text
        if not input_file or not os.path.exists(input_file):
            return jsonify({'status': 'error', 'message': 'Excel file not found after regeneration'})

        wb = load_workbook(input_file)
        sheet_name = get_sheet_name()
        if sheet_name not in wb.sheetnames:
            return jsonify({'status': 'error', 'message': f'{sheet_name} sheet not found in Excel file'})
        ws = wb[sheet_name]

        header = next(ws.rows)
        primary_text_col_name = get_column_name('primary_text')
        secondary_text_col_name = get_column_name('secondary_text')
        
        secondary_text_col_idx = 1
        primary_text_col_idx = 0
        
        for idx, cell in enumerate(header):
            if cell.value == secondary_text_col_name:
                secondary_text_col_idx = idx
            elif cell.value == primary_text_col_name:
                primary_text_col_idx = idx

        excel_row = row_idx + 2
        cell_address = f'{get_column_letter(secondary_text_col_idx + 1)}{excel_row}'
        ws[cell_address].value = new_text
        ws[cell_address].fill = PatternFill(fill_type=None)  # Clear existing fill

        wb.save(input_file)

        # 5. Get updated color status and prepare comparison data
        color_status = get_cell_color_status()
        row_approval = color_status.get(excel_row, {'col_b': False, 'col_b_type': None})
        col_b_approved = row_approval['col_b']
        col_b_type = row_approval['col_b_type']

        # Get original text from Column A for comparison
        col_a_cell = ws[f'{get_column_letter(primary_text_col_idx + 1)}{excel_row}']
        col_a_text = str(col_a_cell.value) if col_a_cell.value is not None else ''
        highlighted_a, highlighted_b, status = compare_text(col_a_text, new_text)

        return jsonify({
            'status': 'success',
            'new_text': new_text,
            'highlighted_html': highlighted_b,
            'highlighted_a_html': highlighted_a,
            'diff_status': status,
            'col_b_approved': col_b_approved,
            'col_b_type': col_b_type
        })

    except FileNotFoundError as e:
        return jsonify({'status': 'error', 'message': str(e)})
    except ValueError as e:
        return jsonify({'status': 'error', 'message': str(e)})
    except Exception as e:
        print(f"Error during custom prompt regeneration for row {row_idx}: {type(e).__name__} - {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': f'An unexpected error occurred: {str(e)}'})

@app.route('/regenerate_multiple_with_custom_prompt', methods=['POST'])
def regenerate_multiple_with_custom_prompt():
    row_ids = request.json.get('row_ids', [])
    custom_prompt = request.json.get('prompt', '')
    provider = request.json.get('provider', 'google')

    print(f"Regenerating rows with custom prompt: {row_ids}")
    
    if not row_ids:
        return jsonify({'status': 'error', 'message': 'No row IDs provided'})
    
    if not custom_prompt.strip():
        return jsonify({'status': 'error', 'message': 'Empty prompt provided'})
    
    try:
        import concurrent.futures
        
        input_file = get_input_file_path()
        results = []
        
        if not input_file or not os.path.exists(input_file):
            return jsonify({'status': 'error', 'message': 'Excel file not found'})
        
        # First, generate all the new texts in parallel
        generated_texts = {}
        
        def generate_text_for_row_custom_prompt(row_idx):
            try:
                print(f"Generating text with custom prompt for row: {row_idx}")
                
                # Read row data
                from src.generate_cell import read_row
                arabic_text, col_a_text, col_b_text = read_row(row_idx, input_file)
                
                # Process the custom prompt by replacing placeholders
                processed_prompt = inject_variables(custom_prompt, {
                    "arabic_text": arabic_text,
                    "col_a_text": col_a_text,
                    "col_b_text": col_b_text
                })
                
                # Generate text using the custom prompt
                from src.ai import ask
                new_text = ask(processed_prompt, provider=provider).text.strip()
                
                return {'status': 'success', 'row_idx': row_idx, 'new_text': new_text}
            except Exception as e:
                import traceback
                return {
                    'status': 'error',
                    'row_idx': row_idx,
                    'message': str(e),
                    'traceback': traceback.format_exc()
                }
        
        # Generate all texts in parallel
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(10, len(row_ids))) as executor:
            # Submit all generation tasks
            future_to_row = {executor.submit(generate_text_for_row_custom_prompt, row_idx): row_idx for row_idx in row_ids}
            
            # Collect results as they complete
            for future in concurrent.futures.as_completed(future_to_row):
                row_idx = future_to_row[future]
                try:
                    result = future.result()
                    if result['status'] == 'success':
                        generated_texts[row_idx] = result['new_text']
                    else:
                        results.append(result)  # Store error results
                except Exception as e:
                    results.append({
                        'status': 'error',
                        'row_idx': row_idx,
                        'message': str(e)
                    })
        
        # If there are successful generations, update the Excel file only once
        if generated_texts:
            try:
                # Use the new batch update function
                batch_results = batch_update_excel_cells(input_file, generated_texts)
                results.extend(batch_results)
                
            except Exception as e:
                import traceback
                error_message = f"Error updating Excel file: {str(e)}"
                traceback.print_exc()
                
                # Add error for each row that was not already recorded as an error
                for row_idx in generated_texts.keys():
                    if not any(r.get('row_idx') == row_idx and r.get('status') == 'error' for r in results):
                        results.append({
                            'status': 'error',
                            'row_idx': row_idx,
                            'message': error_message
                        })
        
        # Return all results
        return jsonify({
            'status': 'success',
            'message': f'Completed regeneration with custom prompt for {len(generated_texts)} rows. {len(row_ids) - len(generated_texts)} failed.',
            'results': results
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': f'An unexpected error occurred: {str(e)}'
        })

def get_all_comments():
    """Get all unique comments from the Excel file for filtering"""
    input_file = get_input_file_path()
    if not input_file or not os.path.exists(input_file):
        return []
    
    try:
        sheet_name = get_sheet_name()
        df = pd.read_excel(input_file, sheet_name=sheet_name)
        
        if 'comments' not in df.columns:
            return []
        
        # Get all non-null comments
        comments = df['comments'].dropna().astype(str)
        
        # Remove empty strings and whitespace-only strings
        comments = comments[comments.str.strip() != '']
        
        # Get unique comments (case-insensitive by converting to lowercase for comparison)
        unique_comments = []
        seen_lower = set()
        
        for comment in comments:
            comment_lower = comment.lower().strip()
            if comment_lower not in seen_lower:
                seen_lower.add(comment_lower)
                unique_comments.append(comment.strip())
        
        # Sort alphabetically (case-insensitive)
        unique_comments.sort(key=str.lower)
        
        return unique_comments
    except Exception as e:
        print(f"Error getting comments: {e}")
        return []

@app.route('/get_all_comments', methods=['GET'])
def get_all_comments_route():
    """API endpoint to get all unique comments for the filter dropdown"""
    try:
        comments = get_all_comments()
        return jsonify({'status': 'success', 'comments': comments})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


# ============================================
# HEALTH CHECK ENDPOINT
# ============================================
@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint for deployment monitoring."""
    return jsonify({'status': 'healthy', 'timestamp': datetime.utcnow().isoformat()})


# ============================================
# FILE UPLOAD ENDPOINTS
# ============================================
@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Upload an Excel file for comparison."""
    try:
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file provided'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400

        if not allowed_file(file.filename):
            return jsonify({'status': 'error', 'message': 'Invalid file type. Only .xlsx and .xls allowed'}), 400

        # Save file
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)

        # Get file info (sheets and columns)
        try:
            xl = pd.ExcelFile(filepath)
            sheets = xl.sheet_names
            # Get columns from first sheet
            df = pd.read_excel(filepath, sheet_name=sheets[0], nrows=0)
            columns = df.columns.tolist()
            xl.close()  # Close the file handle to prevent WinError 32
        except Exception as e:
            columns = []
            sheets = []

        return jsonify({
            'status': 'success',
            'message': 'File uploaded successfully',
            'filename': unique_filename,
            'filepath': filepath,
            'sheets': sheets,
            'columns': columns
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/files', methods=['GET'])
def list_uploaded_files():
    """List all uploaded files."""
    global current_chunk
    try:
        files = []
        upload_folder = app.config['UPLOAD_FOLDER']

        # Get the current active file name (basename only)
        current_file = None
        if current_chunk:
            current_file = os.path.basename(current_chunk)

        if os.path.exists(upload_folder):
            for filename in os.listdir(upload_folder):
                if allowed_file(filename):
                    filepath = os.path.join(upload_folder, filename)
                    stat = os.stat(filepath)
                    files.append({
                        'filename': filename,
                        'size': stat.st_size,
                        'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
                    })

        # Sort by modified date (newest first)
        files.sort(key=lambda x: x['modified'], reverse=True)

        return jsonify({'status': 'success', 'files': files, 'current_file': current_file})

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/files/<filename>/columns', methods=['GET'])
def get_file_columns(filename):
    """Get columns and sheets from an uploaded file."""
    try:
        # Validate filename to prevent path traversal
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'status': 'error', 'message': 'Invalid filename'}), 400

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        # Ensure the resolved path is within the upload folder
        if not os.path.abspath(filepath).startswith(os.path.abspath(app.config['UPLOAD_FOLDER'])):
            return jsonify({'status': 'error', 'message': 'Invalid file path'}), 400

        if not os.path.exists(filepath):
            return jsonify({'status': 'error', 'message': 'File not found'}), 404

        xl = pd.ExcelFile(filepath)
        sheets = xl.sheet_names

        sheet_name = request.args.get('sheet', sheets[0])
        df = pd.read_excel(filepath, sheet_name=sheet_name, nrows=0)
        columns = df.columns.tolist()
        xl.close()  # Close the file handle to prevent WinError 32

        return jsonify({
            'status': 'success',
            'sheets': sheets,
            'columns': columns,
            'current_sheet': sheet_name
        })

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/columns/preview', methods=['POST'])
def preview_column():
    """Get preview data (first 3 rows) from a specific column."""
    try:
        data = request.get_json()
        filename = data.get('filename')
        sheet = data.get('sheet')
        column = data.get('column')

        if not all([filename, sheet, column]):
            return jsonify({'status': 'error', 'message': 'Missing required parameters'}), 400

        # Validate filename to prevent path traversal
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'status': 'error', 'message': 'Invalid filename'}), 400

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        # Ensure the resolved path is within the upload folder
        if not os.path.abspath(filepath).startswith(os.path.abspath(app.config['UPLOAD_FOLDER'])):
            return jsonify({'status': 'error', 'message': 'Invalid file path'}), 400

        if not os.path.exists(filepath):
            return jsonify({'status': 'error', 'message': 'File not found'}), 404

        # Read only the specified column, first 3 rows
        df = pd.read_excel(filepath, sheet_name=sheet, usecols=[column], nrows=3)

        # Convert to list, handling NaN values
        preview_rows = []
        for val in df[column].tolist():
            if pd.isna(val):
                preview_rows.append('')
            else:
                # Truncate long values for preview
                str_val = str(val)
                if len(str_val) > 100:
                    str_val = str_val[:100] + '...'
                preview_rows.append(str_val)

        return jsonify({
            'status': 'success',
            'column': column,
            'rows': preview_rows
        })

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/files/<filename>', methods=['DELETE'])
def delete_file(filename):
    """Delete an uploaded file."""
    try:
        # Validate filename to prevent path traversal (don't use secure_filename as it strips leading underscores)
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'status': 'error', 'message': 'Invalid filename'}), 400

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        # Ensure the resolved path is within the upload folder
        if not os.path.abspath(filepath).startswith(os.path.abspath(app.config['UPLOAD_FOLDER'])):
            return jsonify({'status': 'error', 'message': 'Invalid file path'}), 400

        if os.path.exists(filepath):
            os.remove(filepath)
            return jsonify({'status': 'success', 'message': 'File deleted'})
        else:
            return jsonify({'status': 'error', 'message': 'File not found'}), 404

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/files/select', methods=['POST'])
def select_uploaded_file():
    """Select an uploaded file as the active file."""
    global current_chunk
    try:
        data = request.get_json()
        filename = data.get('filename')

        if not filename:
            return jsonify({'status': 'error', 'message': 'No filename provided'}), 400

        # Validate filename to prevent path traversal
        if '..' in filename or '/' in filename or '\\' in filename:
            return jsonify({'status': 'error', 'message': 'Invalid filename'}), 400

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        # Ensure the resolved path is within the upload folder
        if not os.path.abspath(filepath).startswith(os.path.abspath(app.config['UPLOAD_FOLDER'])):
            return jsonify({'status': 'error', 'message': 'Invalid file path'}), 400

        if not os.path.exists(filepath):
            return jsonify({'status': 'error', 'message': 'File not found'}), 404

        # Set the current chunk to the uploaded file path
        current_chunk = filepath
        print(f"Selected file changed to: {current_chunk}")

        return jsonify({
            'status': 'success',
            'message': f'File "{filename}" is now active',
            'current_file': filename
        })

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/files/deselect', methods=['POST'])
def deselect_file():
    """Deselect the current file (go back to file selection view)."""
    global current_chunk
    try:
        current_chunk = None
        print("File deselected - current_chunk set to None")

        # Clear the cache
        with file_lock:
            excel_cache['df'] = None
            excel_cache['color_status'] = None

        return jsonify({
            'status': 'success',
            'message': 'File deselected successfully'
        })

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ============================================
# GOOGLE SHEETS ENDPOINTS
# ============================================
@app.route('/api/sheets/test', methods=['GET'])
def test_sheets_connection():
    """Test Google Sheets connection."""
    try:
        from src.sheets import test_connection
        result = test_connection()
        return jsonify(result)
    except ImportError:
        return jsonify({'status': 'error', 'message': 'Google Sheets integration not available. Install gspread.'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})


@app.route('/api/sheets/info', methods=['POST'])
def get_sheets_info():
    """Get information about a Google Sheet."""
    try:
        from src.sheets import get_sheet_info
        data = request.get_json()
        url_or_id = data.get('url')

        if not url_or_id:
            return jsonify({'status': 'error', 'message': 'Sheet URL or ID required'}), 400

        info = get_sheet_info(url_or_id)
        return jsonify({'status': 'success', 'info': info})

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/sheets/import', methods=['POST'])
def import_from_sheets():
    """Import data from a Google Sheet."""
    try:
        from src.sheets import import_from_sheets as sheets_import, import_all_worksheets
        data = request.get_json()

        url_or_id = data.get('url')
        worksheet = data.get('worksheet')
        import_all = data.get('import_all', False)

        if not url_or_id:
            return jsonify({'status': 'error', 'message': 'Sheet URL or ID required'}), 400

        uploads_dir = app.config['UPLOAD_FOLDER']

        if import_all:
            # Import all worksheets into a single Excel file
            dataframes, output_path, total_rows = import_all_worksheets(
                url_or_id,
                uploads_dir=uploads_dir
            )

            # Get columns from the first worksheet that has data
            columns = []
            sheets_info = []
            for sheet_name, df in dataframes.items():
                if not columns and len(df.columns) > 0:
                    columns = df.columns.tolist()
                sheets_info.append({'name': sheet_name, 'rows': len(df)})

            return jsonify({
                'status': 'success',
                'message': f'Imported {len(dataframes)} worksheets successfully',
                'filepath': output_path,
                'filename': os.path.basename(output_path),
                'columns': columns,
                'rows': total_rows,
                'sheets': sheets_info
            })
        else:
            # Import single worksheet
            df, output_path = sheets_import(url_or_id, worksheet, uploads_dir=uploads_dir)

            # Get columns from imported data
            columns = df.columns.tolist()

            return jsonify({
                'status': 'success',
                'message': 'Sheet imported successfully',
                'filepath': output_path,
                'filename': os.path.basename(output_path),
                'columns': columns,
                'rows': len(df)
            })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/sheets/export', methods=['POST'])
def export_to_sheets():
    """Export current data to a Google Sheet."""
    try:
        from src.sheets import sync_excel_to_sheets
        data = request.get_json()

        url_or_id = data.get('url')
        excel_path = data.get('excel_path') or get_input_file_path()
        worksheet = data.get('worksheet')

        if not url_or_id:
            return jsonify({'status': 'error', 'message': 'Sheet URL or ID required'}), 400

        result = sync_excel_to_sheets(excel_path, url_or_id, worksheet_name=worksheet)
        return jsonify({'status': 'success', **result})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ============================================
# SETTINGS API ENDPOINTS
# ============================================
@app.route('/api/settings', methods=['GET'])
def get_settings():
    """Get all settings."""
    try:
        from src.models import Settings, ApiKey
        from src.database import decrypt_api_key

        # Get processing settings
        processing_settings = Settings.get_all()

        # Get API keys (masked)
        api_keys = {}
        for api_key in ApiKey.query.all():
            api_keys[api_key.provider] = {
                'model_name': api_key.model_name,
                'max_tokens': api_key.max_tokens,
                'is_active': api_key.is_active,
                'has_key': bool(api_key.api_key_encrypted),
                'api_key_masked': '***' + api_key.api_key_encrypted[-8:] if api_key.api_key_encrypted and len(api_key.api_key_encrypted) > 8 else '***'
            }

        return jsonify({
            'status': 'success',
            'processing': processing_settings,
            'api_keys': api_keys
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/settings/processing', methods=['POST'])
def update_processing_settings():
    """Update processing settings."""
    try:
        from src.models import Settings, db

        data = request.get_json()

        for key, value in data.items():
            if key in Settings.DEFAULTS:
                Settings.set(key, value)

        return jsonify({'status': 'success', 'message': 'Settings updated'})

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/settings/columns', methods=['GET'])
def get_column_settings():
    """Get current column configuration."""
    try:
        columns = {
            'primary_text': get_column_name('primary_text'),
            'secondary_text': get_column_name('secondary_text'),
            'arabic_text': get_column_name('arabic_text'),
            'number': get_column_name('number'),
            'ratio': get_column_name('ratio')
        }
        sheet_name = get_sheet_name()

        return jsonify({
            'status': 'success',
            'columns': columns,
            'sheet_name': sheet_name
        })

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/settings/columns', methods=['POST'])
def update_column_settings():
    """Update column configuration."""
    global config
    try:
        import yaml

        data = request.get_json()
        columns = data.get('columns', {})
        sheet_name = data.get('sheet_name')

        # Load current config file
        config_path = 'config_flash.yaml'
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                config_data = yaml.safe_load(f) or {}
        else:
            config_data = {}

        # Update excel_settings
        if 'excel_settings' not in config_data:
            config_data['excel_settings'] = {}

        if sheet_name:
            config_data['excel_settings']['sheet_name'] = sheet_name

        if 'columns' not in config_data['excel_settings']:
            config_data['excel_settings']['columns'] = {}

        # Update columns
        for key, value in columns.items():
            if value:  # Only set if value is provided
                config_data['excel_settings']['columns'][key] = value

        # Save config file
        with open(config_path, 'w', encoding='utf-8') as f:
            yaml.dump(config_data, f, default_flow_style=False, allow_unicode=True)

        # Reload config
        reload_config()

        # Clear the cache so data is reloaded with new settings
        with file_lock:
            excel_cache['df'] = None
            excel_cache['color_status'] = None
            excel_cache['sheet_name'] = None

        return jsonify({'status': 'success', 'message': 'Column settings updated'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/settings/api-key/<provider>', methods=['POST'])
def update_api_key(provider):
    """Update API key for a provider."""
    try:
        from src.models import ApiKey, db
        from src.database import encrypt_api_key

        valid_providers = ['google', 'claude', 'openai', 'deepseek', 'grok']
        if provider not in valid_providers:
            return jsonify({'status': 'error', 'message': f'Invalid provider. Must be one of: {valid_providers}'}), 400

        data = request.get_json()
        api_key_raw = data.get('api_key', '')
        model_name = data.get('model_name')
        max_tokens = data.get('max_tokens')
        is_active = data.get('is_active', True)

        # Find or create API key entry
        api_key_entry = ApiKey.query.filter_by(provider=provider).first()

        if api_key_entry:
            # Update existing
            if api_key_raw:  # Only update if new key provided
                api_key_entry.api_key_encrypted = encrypt_api_key(api_key_raw)
            if model_name is not None:
                api_key_entry.model_name = model_name
            if max_tokens is not None:
                api_key_entry.max_tokens = max_tokens
            api_key_entry.is_active = is_active
        else:
            # Create new
            if not api_key_raw:
                return jsonify({'status': 'error', 'message': 'API key required for new entry'}), 400

            api_key_entry = ApiKey(
                provider=provider,
                api_key_encrypted=encrypt_api_key(api_key_raw),
                model_name=model_name,
                max_tokens=max_tokens or 4096,
                is_active=is_active
            )
            db.session.add(api_key_entry)

        db.session.commit()

        return jsonify({'status': 'success', 'message': f'API key for {provider} updated'})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/settings/api-key/<provider>/test', methods=['POST'])
def test_api_key(provider):
    """Test API key for a provider."""
    try:
        from src.models import ApiKey
        from src.database import decrypt_api_key

        api_key_entry = ApiKey.query.filter_by(provider=provider, is_active=True).first()

        if not api_key_entry or not api_key_entry.api_key_encrypted:
            return jsonify({'status': 'error', 'message': f'No API key configured for {provider}'}), 404

        api_key = decrypt_api_key(api_key_entry.api_key_encrypted)

        # Test the connection based on provider
        if provider == 'google':
            import google.generativeai as genai
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel(api_key_entry.model_name or 'gemini-2.0-flash')
            response = model.generate_content("Say 'test successful' in 3 words")

        elif provider == 'claude':
            import anthropic
            client = anthropic.Anthropic(api_key=api_key)
            response = client.messages.create(
                model=api_key_entry.model_name or 'claude-3-haiku-20240307',
                max_tokens=50,
                messages=[{"role": "user", "content": "Say 'test successful' in 3 words"}]
            )

        elif provider in ['openai', 'deepseek', 'grok']:
            import openai
            base_urls = {
                'openai': None,
                'deepseek': 'https://api.deepseek.com/v1',
                'grok': 'https://api.x.ai/v1'
            }
            client = openai.OpenAI(api_key=api_key, base_url=base_urls.get(provider))
            response = client.chat.completions.create(
                model=api_key_entry.model_name or 'gpt-4o',
                messages=[{"role": "user", "content": "Say 'test successful' in 3 words"}],
                max_tokens=50
            )

        return jsonify({'status': 'success', 'message': f'API key for {provider} is valid'})

    except Exception as e:
        return jsonify({'status': 'error', 'message': f'API key test failed: {str(e)}'}), 500


# ============================================
# PROJECT MANAGEMENT ENDPOINTS
# ============================================
@app.route('/api/projects', methods=['GET'])
def list_projects():
    """List all projects."""
    try:
        from src.models import Project
        projects = Project.query.order_by(Project.updated_at.desc()).all()
        return jsonify({
            'status': 'success',
            'projects': [p.to_dict() for p in projects]
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/projects', methods=['POST'])
def create_project():
    """Create a new project."""
    try:
        from src.models import Project, db

        data = request.get_json()

        project = Project(
            name=data.get('name', 'Untitled Project'),
            source_type=data.get('source_type', 'upload'),
            source_ref=data.get('source_ref'),
            excel_path=data.get('excel_path'),
            sheet_name=data.get('sheet_name'),
            col_primary_text=data.get('col_primary_text'),
            col_secondary_text=data.get('col_secondary_text'),
            col_arabic_text=data.get('col_arabic_text'),
            col_id=data.get('col_id'),
            col_ratio=data.get('col_ratio'),
            rows_per_chunk=data.get('rows_per_chunk', 500)
        )

        db.session.add(project)
        db.session.commit()

        return jsonify({
            'status': 'success',
            'message': 'Project created',
            'project': project.to_dict()
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/projects/<int:project_id>', methods=['GET'])
def get_project(project_id):
    """Get a specific project."""
    try:
        from src.models import Project
        project = Project.query.get(project_id)

        if not project:
            return jsonify({'status': 'error', 'message': 'Project not found'}), 404

        return jsonify({'status': 'success', 'project': project.to_dict()})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/projects/<int:project_id>', methods=['PUT'])
def update_project(project_id):
    """Update a project."""
    try:
        from src.models import Project, db

        project = Project.query.get(project_id)
        if not project:
            return jsonify({'status': 'error', 'message': 'Project not found'}), 404

        data = request.get_json()

        # Update fields if provided
        for field in ['name', 'source_type', 'source_ref', 'excel_path', 'sheet_name',
                      'col_primary_text', 'col_secondary_text', 'col_arabic_text',
                      'col_id', 'col_ratio', 'rows_per_chunk']:
            if field in data:
                setattr(project, field, data[field])

        db.session.commit()

        return jsonify({
            'status': 'success',
            'message': 'Project updated',
            'project': project.to_dict()
        })

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
def delete_project(project_id):
    """Delete a project."""
    try:
        from src.models import Project, db

        project = Project.query.get(project_id)
        if not project:
            return jsonify({'status': 'error', 'message': 'Project not found'}), 404

        db.session.delete(project)
        db.session.commit()

        return jsonify({'status': 'success', 'message': 'Project deleted'})

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


if __name__ == '__main__':
    host = ServerConfig.get_host()
    port = ServerConfig.get_port()
    debug = not ServerConfig.is_production()
    app.run(host=host, port=port, debug=debug)
