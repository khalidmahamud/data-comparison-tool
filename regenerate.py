import os
import pandas as pd
import google.generativeai as genai
import yaml
import json
import re
import math
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from openpyxl import load_workbook

class GeminiAPI:
    def __init__(self, api_key: str, model: str, name: str):
        genai.configure(api_key=api_key)
        self.client = genai.GenerativeModel(model)
        self.name = name
    
    def generate(self, arabic_text: str, hadith_details: str, prompts: List[str], max_tokens: int) -> List[str]:
        combined_prompt = "\n\n".join([f"Task {i+1}: {prompt}" for i, prompt in enumerate(prompts)])
        combined_text = f"Arabic Text: {arabic_text}\nBangla Translation: {hadith_details}"
        
        try:
            response = self.client.generate_content(
                f"{combined_prompt}\n\nText to process:\n{combined_text}",
                generation_config={"max_output_tokens": max_tokens}
            ).text.strip()
            
            
            return self._parse_response(response, len(prompts))
        except Exception as e:
            print(f"Error in generate_content: {str(e)}")
            return ["Error: " + str(e)] + [""] * (len(prompts) - 1)
        
    def _parse_response(self, response: str, num_prompts: int) -> List[str]:
        if not response:
            print("Warning: Empty response from API")
            return [""] * num_prompts
            
        results = [""] * num_prompts
        
        # Skip JSON parsing and use text parsing directly
        try:
            current_task = 0
            current_content = []
            for line in response.split('\n'):
                task_marker = None
                for i in range(num_prompts):
                    if f"Task {i+1}" in line or f"TASK {i+1}" in line:
                        task_marker = i
                        break
                if task_marker is not None:
                    if current_task > 0 and current_content:
                        results[current_task-1] = '\n'.join(current_content).strip()
                    current_task = task_marker + 1
                    current_content = []
                elif current_task > 0:
                    current_content.append(line)
            if current_task > 0 and current_content:
                results[current_task-1] = '\n'.join(current_content).strip()
                
            # If we still don't have results, split response evenly
            if all(not result for result in results):
                lines = response.split('\n')
                if lines:
                    chunk_size = max(1, len(lines) // num_prompts)
                    for i in range(num_prompts):
                        start = i * chunk_size
                        end = start + chunk_size if i < num_prompts - 1 else len(lines)
                        if start < len(lines):
                            results[i] = '\n'.join(lines[start:min(end, len(lines))]).strip()
        except Exception as e:
            print(f"Error in text parsing: {str(e)}")
            # Last resort - just use the whole response for the first task
            if not any(results):
                results[0] = response
                
        return results

class TextRegenerator:
    def __init__(self, config_path: str = 'config_flash.yaml'):
        self.config = self._load_config(config_path)
        if not self.config:
            raise ValueError("Failed to load configuration")
        self.prompts = self._load_prompts()
        self.api = self._init_api()
    
    def _load_config(self, config_path: str) -> Optional[dict]:
        config_path = Path(config_path)
        if not config_path.exists():
            return None
        try:
            with config_path.open('r') as f:
                return yaml.safe_load(f)
        except Exception:
            return None
    
    def _load_prompts(self) -> List[str]:
        prompts_file = Path(self.config['file_settings']['prompts_file'])
        if not prompts_file.exists():
            raise FileNotFoundError(f"Prompts file '{prompts_file}' not found")
        
        with prompts_file.open('r', encoding='utf-8') as f:
            content = f.read()
            prompts = [prompt.strip() for prompt in content.split('\n\n') if prompt.strip()]
        
        if not prompts:
            raise ValueError(f"Prompts file '{prompts_file}' is empty")
        
        return prompts[:3]  # Only use the first 3 prompts as in the original code
    
    def _init_api(self):
        # Just use the first API config for simplicity
        for name, cfg in self.config['api_settings'].items():
            if name.startswith('gemini_flash'):
                return GeminiAPI(cfg['api_key'], cfg['model'], name)
        raise ValueError("No valid Gemini API configuration found")
    
    def regenerate_text(self, row_idx: int) -> Tuple[str, str]:
        """Regenerate text for a specific row in the Excel file"""
        # Get the Excel file path
        excel_path = Path(self.config['file_settings']['input_file'])
        if not excel_path.exists():
            raise FileNotFoundError(f"Input file '{excel_path}' not found")
        
        # Load the Excel file
        try:
            df = pd.read_excel(excel_path, sheet_name='hadith')
            
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {str(e)}")
        
        # Validate row index
        if row_idx < 0 or row_idx >= len(df):
            raise ValueError(f"Row index {row_idx} out of bounds (0-{len(df)-1})")
        
        # Get hadith_details for the row - we'll use this for both inputs if arabic_text isn't available
        hadith_details = df.loc[row_idx, 'hadith_details'] if 'hadith_details' in df.columns else ""
        
        
        # Get arabic_text if available, otherwise use hadith_details as both inputs
        arabic_text = ""
        if 'arabic_text' in df.columns:
            arabic_text = df.loc[row_idx, 'arabic_text']
        else:
            print("arabic_text column not found, using hadith_details for both inputs")
            arabic_text = hadith_details
        
        # Generate new text
        max_tokens = self.config['api_settings'][self.api.name]['max_tokens']
        results = self.api.generate(arabic_text, hadith_details, self.prompts, max_tokens)
        
        # First result is typically the main correction
        new_text = results[0] if results else ""
        
        # Use openpyxl to update the Excel file (preserves formatting)
        try:
            # Load the workbook with openpyxl
            wb = load_workbook(excel_path)
            if 'hadith' not in wb.sheetnames:
                raise ValueError("'hadith' sheet not found in Excel file")
            
            ws = wb['hadith']
            
            # Find the analysis-3 column index
            analysis3_col_idx = None
            for idx, cell in enumerate(next(ws.rows)):
                if cell.value == 'analysis-3':
                    analysis3_col_idx = idx
                    break
            
            if analysis3_col_idx is None:
                # If column not found, default to column G (index 6)
                analysis3_col_idx = 6  # Column G is index 6 (A=0, B=1, ..., G=6)
                
            
            # Calculate the Excel row (DataFrame index row_idx corresponds to Excel row = row_idx + 2)
            excel_row = row_idx + 2  # +1 for 0-indexed to 1-indexed, +1 for header row
            
            # Get the cell address
            col_letter = chr(65 + analysis3_col_idx)  # Convert to letter (A=0, B=1, etc.)
            cell_address = f'{col_letter}{excel_row}'
            
            
            
            # Update the cell value
            ws[cell_address].value = new_text
            
            # Save the workbook
            wb.save(excel_path)
            print(f"Updated")
            
            return new_text, self.api.name
        
        except Exception as e:
            print(f"Error updating Excel with openpyxl: {str(e)}")
            
            # Fallback to pandas if openpyxl fails
            try:
                print("Falling back to pandas DataFrame update")
                
                # Try to find analysis-3 column in DataFrame
                if 'analysis-3' in df.columns:
                    df.loc[row_idx, 'analysis-3'] = new_text
                    print("Updated 'analysis-3' column in DataFrame")
                else:
                    # If analysis-3 column not found, use column index 6 (column G)
                    if len(df.columns) > 6:
                        df.iloc[row_idx, 6] = new_text  # Update column G (index 6)
                        print("Updated column G (index 6) in DataFrame")
                    else:
                        # If DataFrame doesn't have enough columns, add columns until we reach G
                        while len(df.columns) < 7:
                            col_name = f"Column_{len(df.columns)}"
                            df[col_name] = ""
                        df.iloc[row_idx, 6] = new_text  # Now update column G
                        print(f"Added columns and updated column G (index 6) in DataFrame")
                
                # Save the DataFrame
                output_path = Path(self.config['file_settings']['output_file'])
                df.to_excel(output_path, sheet_name='hadith', index=False)
                print(f"Updated")
            except Exception as e2:
                print(f"Pandas fallback also failed: {str(e2)}")
                # Return the text anyway so the UI can update
            return new_text, self.api.name 